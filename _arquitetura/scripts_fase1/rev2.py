import os,re,io,sys
sys.stdout=io.TextIOWrapper(sys.stdout.buffer,encoding="utf-8")
from openpyxl import load_workbook
from oletools.olevba import VBA_Parser
base=r"C:\Users\vitor\OneDrive - MSFT\Desktop\QC_INI"
f=os.path.join(base,"QC_Hematologia.xlsm")
OUT=open("revdump.txt","w",encoding="utf-8")
def o(*a): OUT.write(" ".join(str(x) for x in a)+"\n")

wb=load_workbook(f,keep_vba=True,data_only=False)
o("=== ABAS ===")
for ws in wb.worksheets:
    o(f"  {ws.title:<16} state={ws.sheet_state:<10} dims={ws.dimensions:<14} prot={ws.protection.sheet}")
o("\n=== NOMES DEFINIDOS (%d) ===" % len(wb.defined_names))
for n,dn in sorted(wb.defined_names.items()):
    o(f"  {n:<16} -> {dn.attr_text}")

# amostra de formulas-chave por aba
o("\n=== FORMULAS-CHAVE ===")
def samp(sheet,cells):
    ws=wb[sheet]
    for c in cells:
        v=ws[c].value
        if v is not None: o(f"  {sheet}!{c}: {str(v)[:230]}")
samp("Resultados",["A4","D4","E4","BA4","BB4","BD2"])
samp("Calc",["A3","B3","C3","D3","E3","F3","G3"])
samp("Liberação",["A4","B4"])
samp("Estatística",["B3","D3","B4","C7","D7","E7","P7","J7","L7"])
samp("Painel",["B3","C3","C4","B7","O3"])
samp("Analitos",["A2","R4","Y4","AA4"])
samp("Configuração",["C20","C21","C26"])
samp("Registros",["A2"])

o("\n=== VBA ===")
p=VBA_Parser(f); os.makedirs("recovered",exist_ok=True)
for (fn_,stream,vba_fname,code) in p.extract_macros():
    safe=vba_fname.replace("/","_")
    open(os.path.join("recovered",safe),"w",encoding="utf-8").write(code)
    procs=re.findall(r'^\s*(?:Public |Private )?(Sub|Function)\s+(\w+)',code,re.M)
    o(f"  {vba_fname} ({len(code.splitlines())} linhas)")
    for k,nm in procs: o(f"      {k} {nm}")
OUT.close()
print(open("revdump.txt",encoding="utf-8").read()[:5200])
