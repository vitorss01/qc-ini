# -*- coding: utf-8 -*-
"""conferir_contrato_modelo.py - o modelo semantico bate com o contrato?

Coluna declarada no modelo que a fonte nao entrega QUEBRA a atualizacao do
Power BI. Coluna que a fonte entrega e o modelo ignora nunca chega a pagina
nenhuma. As duas falham em silencio ate alguem abrir o Desktop.

A comparacao ingenua -- nomes do TMDL contra o cabecalho do BI_Data -- produz
falso positivo, porque entre um e outro existe o Power Query:

  Table.RenameColumns   muda o nome (Sigma -> nada; W_2_2s -> W_R2)
  Table.AddColumn       cria coluna que a fonte nao tem (Nivel_Rotulo)
  sourceColumn:         o TMDL renomeia de novo (Sigma -> Sigma_Obs)
  column X = <DAX>      coluna calculada, sem fonte

Aqui os quatro sao considerados. So sobra o que e divergencia de verdade.
"""
import io
import os
import re
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', write_through=True)
import openpyxl

BASE = r"C:\Users\vitor\OneDrive - MSFT\Desktop\QC_INI"
MOD = os.path.join(BASE, 'PowerBI', 'QC_INI_Bioquimica.SemanticModel', 'definition')
FATO = os.path.join(MOD, 'tables', 'Fato_QC.tmdl')

ok = fail = 0


def reg(nome, cond, detalhe=''):
    global ok, fail
    if cond:
        ok += 1
    else:
        fail += 1
    print('  %-5s %-52s %s' % ('OK' if cond else 'FALHA', nome[:52], detalhe[:60]))


def ler(p):
    raw = io.open(p, 'rb').read()
    enc = 'utf-8-sig' if raw[:3] == b'\xef\xbb\xbf' else 'utf-8'
    return raw.decode(enc)


def cab_bi(arq):
    wb = openpyxl.load_workbook(arq, read_only=True)
    if 'BI_Data' not in wb.sheetnames:
        wb.close()
        return []
    ws = wb['BI_Data']
    c = [str(x.value or '') for x in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.close()
    return [x for x in c if x]


def main():
    s = ler(FATO)

    # colunas do TMDL: calculadas x vindas de fonte
    calc = set(re.findall(r'^\s*column\s+(\w+)\s*=', s, re.M))
    src = {}
    for m in re.finditer(r'^\s*column\s+([\w\']+)\s*$(.*?)(?=^\s*(?:column|measure|partition|hierarchy)\s|\Z)',
                         s, re.M | re.S):
        nome = m.group(1).strip("'")
        sc = re.search(r'sourceColumn:\s*(\S+)', m.group(2))
        if sc:
            src[nome] = sc.group(1)

    # o que o M cria e o que ele renomeia
    add = set(re.findall(r'Table\.AddColumn\([^,]+,\s*"([^"]+)"', s))
    ren = {}
    for m in re.finditer(r'\{"([^"]+)"\s*,\s*"([^"]+)"\}', s):
        a, b = m.group(1), m.group(2)
        if a != b and not re.match(r'^(type|Int64|true|false)', b):
            ren[a] = b

    bio = set(cab_bi(os.path.join(BASE, 'QC_Bioquimica.xlsm')))
    hem = set(cab_bi(os.path.join(BASE, 'QC_Hematologia.xlsm')))

    def efetivas(fonte):
        """nomes disponiveis depois do Power Query"""
        e = set()
        for c in fonte:
            e.add(ren.get(c, c))
        return e | add

    ebio, ehem = efetivas(bio), efetivas(hem)

    print('modelo: %d colunas de fonte, %d calculadas, %d medidas'
          % (len(src), len(calc), len(re.findall(r'^\s*measure\s', s, re.M))))
    print('BI_Data: Bioquimica %d, Hematologia %d colunas' % (len(bio), len(hem)))
    print()

    print('=== toda coluna do modelo tem origem? ===')
    faltam_bio = sorted(v for k, v in src.items() if v not in ebio)
    faltam_hem = sorted(v for k, v in src.items() if v not in ehem)
    reg('modelo x Bioquimica', not faltam_bio,
        'faltam: %s' % (faltam_bio[:4] or 'nada'))
    reg('modelo x Hematologia', not faltam_hem,
        'faltam: %s' % (faltam_hem[:4] or 'nada'))

    print()
    print('=== os dois produtos entregam o MESMO esquema apos o rename? ===')
    dif = sorted(ebio ^ ehem)
    reg('esquema identico (Table.Combine seguro)', not dif,
        'divergem: %s' % (dif[:5] or 'nenhuma'))

    print()
    print('=== a escada de Sigma e unica? ===')
    escadas = re.findall(r'column\s+\w+\s*=\s*SWITCH\s*\(\s*TRUE\s*\(\)[^\n]*Sigma_Obs[^\n]*', s)
    reg('nenhuma escada de Sigma reimplementada em DAX', not escadas,
        '%d encontrada(s)' % len(escadas))
    usa_contrato = 'Fato_QC[Classificacao_Sigma]' in s
    reg('classificacao vem do contrato', usa_contrato)

    print()
    print('=== nomes de Westgard ===')
    reg('W_10x eliminado do modelo', 'W_10x' not in s)
    for p in ('W_R2', 'W_R4', 'W_R5'):
        reg('posicao divergente com nome neutro: %s' % p, p in s)
    dim = os.path.join(MOD, 'tables', 'Dim_Regra_Westgard.tmdl')
    reg('Dim_Regra_Westgard existe', os.path.exists(dim))
    mt = ler(os.path.join(MOD, 'model.tmdl'))
    reg('dimensao declarada no model.tmdl', 'ref table Dim_Regra_Westgard' in mt)
    if os.path.exists(dim):
        d = ler(dim)
        reg('dimensao traz 2of3_2s / 3_1s / 6x da Hematologia',
            all(x in d for x in ('2of3_2s', '3_1s', '6x')))
        reg('dimensao traz 2_2s / 4_1s / 8x da Bioquimica',
            all(('"%s"' % x) in d for x in ('2_2s', '4_1s', '8x')))
        # o comentario EXPLICA por que 10x nao existe mais; o que nao pode e
        # 10x aparecer como DADO, numa das dez linhas da tabela.
        dados = '\n'.join(l for l in d.split('\n') if not l.strip().startswith('///'))
        reg('10x nao aparece como dado da dimensao', '10x' not in dados)

    print()
    print('=== Eventos_Westgard como fato de EVENTO ===')
    fev = os.path.join(MOD, 'tables', 'Fato_Eventos_Westgard.tmdl')
    reg('Fato_Eventos_Westgard existe', os.path.exists(fev))
    reg('registrada no model.tmdl', 'ref table Fato_Eventos_Westgard' in mt)
    if os.path.exists(fev):
        e = ler(fev)
        for m in ('N_Eventos_Violacao', 'N_Resultados_Marcados',
                  'N_Corridas_Com_Violacao'):
            reg('medida %s' % m, ('measure %s' % m) in e)
        reg('separa evento de resultado marcado (N x R)',
            'N_Niveis' in e and 'R_Corridas' in e)
        reg('declara o grao por produto', 'Grao' in e and 'LEGADO_1N' in e)
        reg('grao legado nao entra na contagem de eventos',
            'Grao] = "EVENTO"' in e)
        reg('expoe quanto ficou fora', 'Eventos_sem_grao_de_evento' in e)
    # as duas abas existem e tem dado
    import openpyxl as _op
    for arq, prod in (('QC_Bioquimica.xlsm', 'Bioquimica'),
                      ('QC_Hematologia.xlsm', 'Hematologia')):
        wb = _op.load_workbook(os.path.join(BASE, arq), read_only=True)
        tem = 'Eventos_Westgard' in wb.sheetnames
        n = 0
        if tem:
            ws = wb['Eventos_Westgard']
            n = max(0, ws.max_row - 3)
        wb.close()
        reg('%s: aba de eventos com dado' % prod, tem and n > 0, '%d linha(s)' % n)

    print()
    print('=== EQA integrada ===')
    feq = os.path.join(MOD, 'tables', 'Fato_EQA.tmdl')
    reg('Fato_EQA existe', os.path.exists(feq))
    reg('registrada no model.tmdl', 'ref table Fato_EQA' in mt)
    if os.path.exists(feq):
        q_ = ler(feq)
        for c in ('Provedor', 'Ano', 'Rodada', 'Analito', 'Area'):
            reg('filtro disponivel: %s' % c, ('column %s' % c) in q_ or
                ('sourceColumn: %s' % c) in q_)
        for m in ('N_EQA_Resultados', 'N_EQA_Rodadas', 'SDI_Medio',
                  'Bias_EQA_pct', 'Bias_Abs_EQA_pct', 'N_Nao_Conformidades',
                  'Aceitabilidade_pct'):
            reg('medida %s' % m, ('measure %s' % m) in q_)
        reg('SDI preservado separado do Bias',
            'column SDI' in q_ and 'column Bias_pct' in q_)
        reg('|Bias| consolidado em dois estagios',
            q_.count('AVERAGEX (') >= 2 and 'VALUES ( Fato_EQA[Rodada] )' in q_)
        reg('aceitabilidade sem resultado devolve BLANK', 'BLANK ( )' in q_ or 'BLANK ()' in q_)
    import openpyxl as _o
    for arq, prod in (('QC_Bioquimica.xlsm', 'Bioquimica'),
                      ('QC_Hematologia.xlsm', 'Hematologia')):
        wb = _o.load_workbook(os.path.join(BASE, arq), read_only=True)
        tem = 'EQA_Base' in wb.sheetnames
        n = 0
        cab = []
        if tem:
            ws = wb['EQA_Base']
            cab = [str(c.value or '') for c in next(ws.iter_rows(min_row=1, max_row=1))]
            n = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0])
        wb.close()
        reg('%s: EQA_Base com dado' % prod, tem and n > 0, '%d linha(s)' % n)
        # as 21 primeiras sao o dado; o que vem depois e bloco lateral
        # (mapa de nomes do provedor e carimbo), e o M recorta por nome.
        ESPERADO = ['Provedor', 'Ano', 'Rodada', 'Analito', 'Analito_Canonico',
                    'Amostra', 'Resultado', 'Valor_Alvo', 'SD', 'SDI',
                    'Limite_Inferior', 'Limite_Superior', 'Avaliacao_Original',
                    'Status_Padronizado', 'Unidade', 'Bias', 'Bias_Abs',
                    'Pagina_Fonte', 'Arquivo_Fonte', 'Uso_Analitico', 'Chave']
        reg('%s: 21 colunas de contrato na ordem' % prod,
            cab[:21] == ESPERADO,
            'difere em %s' % [c for a_, c in zip(ESPERADO, cab[:21]) if a_ != c][:3])

    print()
    print('=== caminhos das fontes ===')
    ex = ler(os.path.join(MOD, 'expressions.tmdl'))
    reg('parametros apontam para o repositorio', 'vitor.santos' not in ex,
        'ainda aponta para vitor.santos' if 'vitor.santos' in ex else '')
    for arq in ('QC_Bioquimica.xlsm', 'QC_Hematologia.xlsm'):
        reg('fonte existe: %s' % arq, os.path.exists(os.path.join(BASE, arq)))

    print()
    print('%d OK, %d FALHA' % (ok, fail))
    sys.exit(1 if fail else 0)


main()
