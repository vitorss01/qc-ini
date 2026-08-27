# -*- coding: utf-8 -*-
"""validar_entrega_fase1.py - a base do Excel entregue esta integra?

Compara o arquivo ATUAL contra o backup tirado ANTES desta intervencao, para
provar que o reparo nao custou dado nem layout:

  dados      DB_Resultados, EQA_Base, LotesStore, Analitos: contagem de linhas
  estrutura  abas antes e depois (as tres novas sao esperadas)
  Painel     valores das celulas que o gestor ajustou a mao (ADR-036)
  VBA        modulos antes e depois (os dois novos sao esperados)
  formulas   nenhuma celula em erro

Nao abre o Excel: le com openpyxl e oletools, entao nao altera nada.
"""
import io
import os
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', write_through=True)
import openpyxl
from oletools.olevba import VBA_Parser

BASE = r"C:\Users\vitor\OneDrive - MSFT\Desktop\QC_INI"
ATUAL = os.path.join(BASE, 'QC_Bioquimica.xlsm')
ANTES = os.path.join(BASE, '_arquitetura', '_backup_producao',
                     'QC_Bioquimica_pre_fase1_abas.xlsm')

ABAS_ESPERADAS = {'Eng_Saida', 'Cfg_Status', 'Eventos_Westgard'}
MODS_ESPERADOS = {'mWestgardKnowledge'}

ok = fail = 0


def reg(nome, esp, obt, cond):
    global ok, fail
    if cond:
        ok += 1
    else:
        fail += 1
    print('  %-5s %-44s esp %-22s obt %s'
          % ('OK' if cond else 'FALHA', nome[:44], str(esp)[:22], str(obt)[:40]))


def linhas(ws, col=1):
    n = 0
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, col).value not in (None, ''):
            n += 1
    return n


def mods(arq):
    p = VBA_Parser(arq)
    s = {m[2].replace('.bas', '').replace('.cls', '').replace('.frm', '')
         for m in p.extract_macros() if m[2]}
    p.close()
    return s


def main():
    for c in (ATUAL, ANTES):
        if not os.path.exists(c):
            raise SystemExit('ausente: %s' % c)

    a = openpyxl.load_workbook(ANTES)
    b = openpyxl.load_workbook(ATUAL)

    print('=== estrutura de abas ===')
    novas = set(b.sheetnames) - set(a.sheetnames)
    sumidas = set(a.sheetnames) - set(b.sheetnames)
    reg('abas novas sao exatamente as tres do motor',
        sorted(ABAS_ESPERADAS), sorted(novas), novas == ABAS_ESPERADAS)
    reg('nenhuma aba sumiu', 'nenhuma', sorted(sumidas) or 'nenhuma', not sumidas)

    print()
    print('=== dados ===')
    for nome, col in (('DB_Resultados', 1), ('EQA_Base', 1),
                      ('LotesStore', 1), ('Analitos', 1)):
        if nome not in a.sheetnames or nome not in b.sheetnames:
            reg('%s presente nos dois' % nome, 'sim', 'nao', False)
            continue
        na, nb = linhas(a[nome], col), linhas(b[nome], col)
        reg('%s: linhas preservadas' % nome, na, nb, na == nb)

    print()
    print('=== Painel: o que o gestor ajustou a mao (ADR-036) ===')
    pa, pb = a['Painel'], b['Painel']
    refs = ['F6', 'G6', 'H6', 'I6', 'J6', 'K6', 'L6', 'M6',
            'R9', 'S9', 'V9', 'W9', 'X9', 'Y9', 'R22', 'R23']
    difs = []
    for r in refs:
        va, vb = pa[r].value, pb[r].value
        if va != vb:
            difs.append('%s: %r -> %r' % (r, va, vb))
    reg('rotulos e ancoras do Painel intactos', 'sem diferenca',
        difs[:2] or 'sem diferenca', not difs)

    larg = [c for c in ('F', 'G', 'H', 'I', 'J', 'K', 'R', 'S', 'V')
            if (pa.column_dimensions[c].width if c in pa.column_dimensions else None)
            != (pb.column_dimensions[c].width if c in pb.column_dimensions else None)]
    reg('larguras de coluna do Painel intactas', 'nenhuma mudou',
        larg or 'nenhuma mudou', not larg)

    mescl_a = {str(m) for m in pa.merged_cells.ranges}
    mescl_b = {str(m) for m in pb.merged_cells.ranges}
    reg('mesclagens do Painel intactas', len(mescl_a), len(mescl_b),
        mescl_a == mescl_b)

    print()
    print('=== VBA ===')
    ma, mb = mods(ANTES), mods(ATUAL)
    novos = mb - ma
    perdidos = ma - mb
    reg('modulo novo e o que faltava', sorted(MODS_ESPERADOS), sorted(novos),
        MODS_ESPERADOS.issubset(novos))
    reg('nenhum modulo perdido', 'nenhum', sorted(perdidos) or 'nenhum',
        not perdidos)

    print()
    print('=== formulas ===')
    ruins = []
    for ws in b.worksheets:
        for linha in ws.iter_rows():
            for cel in linha:
                v = cel.value
                if isinstance(v, str) and v.startswith('#') and v.strip('#'):
                    ruins.append('%s!%s=%s' % (ws.title, cel.coordinate, v))
    reg('nenhuma celula com erro gravado', 0, '%d %s' % (len(ruins), ruins[:3]),
        not ruins)

    a.close()
    b.close()

    print()
    print('%d OK, %d FALHA' % (ok, fail))
    sys.exit(1 if fail else 0)


main()
