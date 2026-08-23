# -*- coding: utf-8 -*-
"""migrar_dados_eqa.py - ADR-034: os dados entram no modulo novo

DUAS ORIGENS, TRATAMENTOS DIFERENTES

  1. CAP_C3_2025_COMPLETO_ORGANIZADO.xlsx -- 455 resultados reais, 31 analitos,
     3 surveys. Entram como dado analitico pleno.

  2. EQC_Dados (aba legada) -- 90 registros que NAO sao resultado real de EP:
     glicose entre 250 e 262 mg/dL nas quinze amostras, com quatro casas
     decimais, e nenhum valor em comum com o relatorio do CAP. E simulacao.
     Entram, ficam visiveis e contados, mas marcados no Arquivo Fonte como
     simulacao -- e o mEQA poe Uso_Analitico = NAO neles.

     Apagar violaria "nao apagar historico". Misturar contaminaria todo Sigma e
     todo ET da pasta. A coluna existe para nao ser preciso escolher.

O MAPA DE ANALITOS

O CAP reporta "Urea Nitrogen"; a pasta chama "Ureia". Sem mapa, nenhuma linha
cruzaria com a Analitos e a Estatistica ficaria vazia. O mapa vai para W:Y da
EQA_Base, editavel, e sai daqui como semente.

DOIS ANALITOS DO CAP FICAM SEM CORRESPONDENTE

Ferritin e Thyroid Stim Hormone nao existem na Bioquimica desta pasta. Entram na
base com Analito_Canonico vazio: visiveis, contados, rastreaveis, e fora da
Estatistica -- porque nao ha especificacao de qualidade para cruzar com eles.
Inventar correspondente seria pior do que a lacuna.

BUN E UREIA NAO SAO A MESMA ESCALA

"Urea Nitrogen" do CAP e nitrogenio ureico; "Ureia" e ureia. Diferem por um
fator de 2,14. O Bias em PORCENTAGEM nao se altera com isso -- resultado e alvo
estao na mesma unidade, e o fator cancela na razao. Ja SD, SDI e limites ficam
na escala do provedor, que e onde eles sao usados. Por isso o mapeamento e
correto e a unidade original fica preservada na coluna N.

Uso: python migrar_dados_eqa.py <arquivo.xlsm> <referencia.xlsx>
"""
import io
import os
import re
import sys
import time
import subprocess

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', write_through=True)
import win32com.client as w

CRLF = chr(13) + chr(10)
import openpyxl

MARCA_SIM = 'EQC_Dados (simulação pré-ADR-034)'

# CAP (como o provedor escreve) -> analito canonico da pasta
MAPA_CAP = {
    'ALT (SGPT)': 'ALT (TGP)',
    'AST (SGOT)': 'AST (TGO)',
    'Albumin': 'Albumina',
    'Alkaline Phosphatase': 'Fosfatase alcalina',
    'Amylase, serum': 'Amilase',
    'Bilirubin, direct': 'Bilirrubina direta',
    'Bilirubin, total': 'Bilirrubina total',
    'CO2': 'Bicarbonato',
    'Calcium, serum': 'Cálcio',
    'Chloride': 'Cloro',
    'Cholesterol': 'Colesterol total',
    'Creatine Kinase': 'Creatina fosfoquinase',
    'Creatinine, serum': 'Creatinina',
    'Ferritin': '',                         # a Bioquimica desta pasta nao mede
    'Gamma Glutamyl Trans': 'GGT',
    'Glucose, serum': 'Glicose',
    'HDL Cholesterol': 'HDL colesterol',
    'Iron': 'Ferro',
    'Lactate': 'Lactato',
    'Lactate Dehydrogenase': 'Lactato desidrogenase',
    'Lipase': 'Lipase',
    'Magnesium': 'Magnésio',
    'Phosphorus, serum': 'Fósforo',
    'Potassium, serum': 'Potássio',
    'Protein, total, serum': 'Proteína total',
    'Sodium, serum': 'Sódio',
    'TIBC': 'Capacidade de fixação do ferro',
    'Thyroid Stim Hormone': '',             # idem
    'Triglycerides': 'Triglicerídeos',
    'Urea Nitrogen': 'Ureia',
    'Uric Acid': 'Ácido úrico',
}



def importar_bas(vbp, caminho):
    """Importa um .bas convertendo o arquivo para cp1252.

    O VBA le .bas como ANSI da pagina de codigo do sistema. Um arquivo gravado
    em UTF-8 chega com cada acento virando dois caracteres: "nao aplicavel"
    escrito com til e acento apareceu na tela do Painel como dois simbolos por
    letra. Converter na hora da importacao resolve para todos os modulos de uma
    vez, e mantem os fontes legiveis em UTF-8 no repositorio.
    """
    import tempfile
    texto = io.open(caminho, encoding='utf-8', errors='replace').read()
    tmp = os.path.join(tempfile.gettempdir(),
                       'ansi_' + os.path.basename(caminho))
    with io.open(tmp, 'w', encoding='cp1252', errors='replace',
                 newline=CRLF) as fh:
        fh.write(texto)
    vbp.VBComponents.Import(tmp)


def novo_excel():
    for t in range(1, 9):
        try:
            xl = w.DispatchEx('Excel.Application')
            xl.Visible = False
            xl.DisplayAlerts = False
            xl.EnableEvents = False
            xl.AutomationSecurity = 1
            return xl
        except Exception:
            if t in (1, 3, 5):
                subprocess.call(['powershell', '-NoProfile', '-Command',
                                 'Start-Process excel.exe -WindowStyle Minimized'],
                                stderr=subprocess.DEVNULL)
                time.sleep(10)
            time.sleep(2.5 * t)
    raise RuntimeError('Excel COM nao subiu')


def tenta(fn, vezes=8):
    ult = None
    for i in range(vezes):
        try:
            return fn()
        except Exception as e:
            ult = e
            s = str(e).lower()
            if 'rejeitada' not in s and 'rejected' not in s:
                raise
            time.sleep(1.0 + 0.8 * i)
    raise ult


def ler_referencia(caminho):
    """Devolve as linhas do CAP na estrutura canonica A..R (sem O/P, que sao
    formula na planilha)."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb['CAP_Evaluation_Data']
    linhas = []
    for r in range(2, ws.max_row + 1):
        survey = ws.cell(r, 1).value
        analito = ws.cell(r, 2).value
        if not analito:
            continue
        m = re.search(r'(\d{4})', str(survey or ''))
        ano = int(m.group(1)) if m else None
        linhas.append([
            'CAP',                      # A provedor
            str(survey).strip(),        # B survey
            ano,                        # C ano
            str(analito).strip(),       # D analito
            ws.cell(r, 3).value,        # E amostra
            ws.cell(r, 4).value,        # F resultado
            ws.cell(r, 5).value,        # G media
            ws.cell(r, 6).value,        # H SD
            ws.cell(r, 7).value,        # I SDI
            ws.cell(r, 8).value,        # J lim inf
            ws.cell(r, 9).value,        # K lim sup
            ws.cell(r, 10).value,       # L avaliacao
            ws.cell(r, 11).value,       # M n labs
            ws.cell(r, 12).value,       # N unidade
            ws.cell(r, 15).value,       # Q pagina
            ws.cell(r, 16).value,       # R arquivo
        ])
    wb.close()
    # A ordem primaria e a RODADA: C-A inteiro, depois C-B, depois C-C.
    # Agrupar por analito misturaria surveys, que e exatamente o que a secao 9
    # da missao proibe.
    linhas.sort(key=lambda x: (x[2] or 0, x[1], str(x[3]), str(x[4])))
    return linhas


def ler_legado(wsEQC):
    """Os 90 registros da aba legada, na mesma estrutura canonica."""
    ult = tenta(lambda: wsEQC.Cells(wsEQC.Rows.Count, 1).End(-4162).Row)
    if ult < 4:
        return []
    d = tenta(lambda: wsEQC.Range(wsEQC.Cells(4, 1), wsEQC.Cells(ult, 15)).Value)
    linhas = []
    for row in d:
        if not row[0] or not str(row[0]).strip():
            continue
        linhas.append([
            str(row[4] or 'CAP').strip(),        # A provedor (col E da legada)
            str(row[2] or '').strip(),           # B rodada  (col C)
            int(row[1]) if row[1] else None,     # C ano     (col B)
            str(row[0]).strip(),                 # D analito (col A) -- ja canonico
            str(row[5] or '').strip(),           # E amostra (col F)
            row[6],                              # F resultado
            row[7],                              # G media
            row[8],                              # H SD
            row[9],                              # I SDI
            row[10],                             # J lim inf
            row[11],                             # K lim sup
            row[12],                             # L avaliacao
            None,                                # M n labs -- a legada nao tem
            None,                                # N unidade -- idem
            None,                                # Q pagina
            MARCA_SIM,                           # R arquivo -> Uso_Analitico=NAO
        ])
    linhas.sort(key=lambda x: (x[2] or 0, x[1], str(x[3]), str(x[4])))
    return linhas


def gravar(ws, lo, linhas, nome):
    """Escreve o bloco e devolve quantas linhas ficaram na tabela."""
    n = len(linhas)
    if n == 0:
        print('%s: nenhuma linha para gravar' % nome)
        return 0

    # a tabela precisa ter o tamanho ANTES da escrita: escrever fora dela
    # deixaria o dado orfao, sem filtro, sem formato e sem coluna calculada
    tenta(lambda: lo.Resize(ws.Range(ws.Cells(1, 1), ws.Cells(1 + n, 21))))

    # A..N e Q..R sao dado; O, P e as tres de apoio sao formula
    bloco = [r[:14] for r in linhas]
    tenta(lambda: ws.Range(ws.Cells(2, 1), ws.Cells(1 + n, 14)).__setattr__('Value', bloco))
    bloco2 = [[r[14], r[15]] for r in linhas]
    tenta(lambda: ws.Range(ws.Cells(2, 17), ws.Cells(1 + n, 18)).__setattr__('Value', bloco2))

    # .Formula numa faixa de varias celulas: o Excel ajusta as referencias
    # relativas sozinho. As ancoras absolutas ($L$2) ficam paradas e as
    # relativas ($L2) acompanham a linha -- que e exatamente o que as colunas
    # de apoio precisam para contar "quantos ate aqui".
    for c, f in ((15, '=IFERROR((F2-G2)/G2*100,"")'),
                 (16, '=IF(O2="","",ABS(O2))'),
                 (19, '=IF($L2="","",IF(mEQA.PadronizarStatus($L2)="NAO ACEITO",'
                      'COUNTIFS($L$2:$L2,$L2),""))'),
                 (20, '=IF($B2="","",IF(COUNTIF($B$2:$B2,$B2)=1,'
                      'SUMPRODUCT((COUNTIF($B$2:$B2,$B$2:$B2)=1)*1),""))'),
                 (21, '=IF($D2="","",IF(COUNTIF($D$2:$D2,$D2)=1,'
                      'SUMPRODUCT((COUNTIF($D$2:$D2,$D$2:$D2)=1)*1),""))')):
        tenta(lambda cc=c, ff=f: ws.Range(ws.Cells(2, cc), ws.Cells(1 + n, cc))
              .__setattr__('Formula', ff))
    print('%s: %d linhas gravadas' % (nome, n))
    return n


def main(caminho, referencia):
    caminho = os.path.abspath(caminho)
    linhasCAP = ler_referencia(referencia)
    print('referencia lida: %d resultados' % len(linhasCAP))

    xl = novo_excel()
    wb = xl.Workbooks.Open(caminho)
    salvou = False
    if wb.ReadOnly:
        wb.Close(False)
        xl.Quit()
        raise SystemExit('Somente leitura: %s' % caminho)
    try:
        xl.Calculation = -4135

        # reimporta o motor: este script roda depois de correcoes no mEQA, e
        # rodar a consolidacao com a versao antiga do modulo mediria a coisa
        # errada
        raiz = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        vbp = wb.VBProject
        for nome in ('mEQA', 'mCEQ'):
            for c in list(vbp.VBComponents):
                if c.Name == nome:
                    vbp.VBComponents.Remove(c)
            importar_bas(vbp, os.path.join(raiz, 'src_producao', nome + '.bas'))
        print('modulos reimportados: mEQA, mCEQ')
        # prova de compilacao antes de qualquer medida
        print('   MapearAnalito("CAP","Urea Nitrogen") = %r'
              % tenta(lambda: xl.Run('MapearAnalito', 'CAP', 'Urea Nitrogen')))

        cap = wb.Worksheets('EQA.CAP_Dados')
        base = wb.Worksheets('EQA_Base')
        eqc = None
        for ws in wb.Worksheets:
            if ws.Name == 'EQC_Dados':
                eqc = ws

        # ---- 1. mapa de analitos na EQA_Base -----------------------------
        mapa = [['CAP', k, v] for k, v in sorted(MAPA_CAP.items())]
        legado = ler_legado(eqc) if eqc is not None else []
        for nome in sorted(set(str(r[3]) for r in legado)):
            # a aba legada ja guardava o nome canonico: mapa identidade, para
            # a linha continuar rastreavel mesmo estando fora da analise
            mapa.append(['CAP', nome, nome])
        tenta(lambda: base.Range(base.Cells(2, 23), base.Cells(201, 25)).ClearContents())
        tenta(lambda: base.Range(base.Cells(2, 23),
                                 base.Cells(1 + len(mapa), 25)).__setattr__('Value', mapa))
        semCanon = sorted(k for k, v in MAPA_CAP.items() if not v)
        print('mapa de analitos: %d entradas (%d sem correspondente: %s)'
              % (len(mapa), len(semCanon), ', '.join(semCanon)))

        # ---- 2. CAP real + legado marcado --------------------------------
        todas = linhasCAP + legado
        lo = cap.ListObjects('tblEQA_CAP_Dados')
        n = gravar(cap, lo, todas, 'EQA.CAP_Dados')
        print('   %d do CAP real + %d de simulacao preservada'
              % (len(linhasCAP), len(legado)))

        # ---- 3. consolidar ------------------------------------------------
        xl.Calculation = -4105
        tenta(lambda: xl.CalculateFullRebuild())
        tenta(lambda: xl.Run('AtualizarEQABase'))
        carimbo = tenta(lambda: base.Cells(1, 26).Value)
        print('EQA_Base -> %s' % carimbo)

        # ---- 4. conferencia ----------------------------------------------
        ultB = tenta(lambda: base.Cells(base.Rows.Count, 1).End(-4162).Row)
        print('EQA_Base: %d linhas' % max(0, ultB - 1))
        erros = 0
        for r in range(2, min(1 + n, 300) + 1):
            for c in (15, 16):
                t = str(tenta(lambda rr=r, cc=c: cap.Cells(rr, cc).Text))
                if t.startswith('#') and t.strip('#') != '':
                    erros += 1
                    if erros <= 3:
                        print('   ERRO %s%d = %s' % ('OP'[c - 15], r, t))
        print('celulas de bias em erro (300 primeiras linhas): %d' % erros)
        if erros:
            raise SystemExit('erro de formula -- nada salvo')

        wb.Save()
        salvou = True
        print('SALVO: %s' % caminho)
    finally:
        try:
            wb.Close(salvou)
        except Exception:
            pass
        try:
            xl.Quit()
        except Exception:
            pass


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
