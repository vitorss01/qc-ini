# -*- coding: utf-8 -*-
"""testar_status_westgard.py - ADR-036: M7/M8 e a cadeia dos filtros de EQA

Trabalha numa COPIA em TEMP.

M7 e M8 sao formula: M7 = f(B7, L7). A forma honesta de prova-las e alimentar
f com entradas controladas e conferir a saida -- e, em separado, mostrar que
B7/L7 vem mesmo do motor do Calc, e nao de Sigma, DPM ou margem de ETp.

Uso: python testar_status_westgard.py <arquivo.xlsm> <referencia.xlsx>
"""
import io
import os
import sys
import time
import shutil
import subprocess

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', write_through=True)
import win32com.client as w
import openpyxl

falhas = []
LIN = 14
ANO_TESTE = 2099
ROD = 'TESTE-2099'


def ck(nome, cond, det=''):
    print(('  OK   ' if cond else '  FALHA') + '  ' + nome +
          (('  -> ' + det) if det else ''))
    if not cond:
        falhas.append(nome)


def txt(v):
    return '' if v is None else str(v).strip()


def perto(a, b, tol=1e-9):
    try:
        return abs(float(a) - float(b)) <= tol
    except Exception:
        return False


def novo_excel():
    for t in range(1, 9):
        try:
            xl = w.DispatchEx('Excel.Application')
            xl.Visible = False
            xl.DisplayAlerts = False
            xl.EnableEvents = False
            xl.AutomationSecurity = 1
            return xl
        except Exception:
            if t in (1, 3, 5):
                subprocess.call(['powershell', '-NoProfile', '-Command',
                                 'Start-Process excel.exe -WindowStyle Minimized'],
                                stderr=subprocess.DEVNULL)
                time.sleep(10)
            time.sleep(2.5 * t)
    raise RuntimeError('Excel COM nao subiu')


TRANS = ('rejeitada', 'rejected', 'membro n', 'member not found', 'busy')


def tenta(fn, vezes=10):
    ult = None
    for i in range(vezes):
        try:
            return fn()
        except Exception as e:
            ult = e
            if not any(t in str(e).lower() for t in TRANS):
                raise
            time.sleep(1.0 + 0.8 * i)
    raise ult


def main(caminho, referencia):
    copia = os.path.join(os.environ.get('TEMP', '.'),
                         'tsw_' + os.path.basename(caminho))
    shutil.copy(caminho, copia)
    xl = novo_excel()
    wb = xl.Workbooks.Open(copia)
    try:
        pa = wb.Worksheets('Painel')
        es = wb.Worksheets('Estatística')
        an = wb.Worksheets('Analitos')
        cap = wb.Worksheets('EQA.CAP_Dados')
        base = wb.Worksheets('EQA_Base')
        base.Visible = -1
        xl.Calculation = -4105

        fM7 = txt(tenta(lambda: pa.Range('M7').Formula))
        fM8 = txt(tenta(lambda: pa.Range('M8').Formula))
        fL7 = txt(tenta(lambda: pa.Range('L7').Formula))
        fG7 = txt(tenta(lambda: pa.Range('G7').Formula))

        print('=== ONDE M7/M8 BEBEM ===')
        print('   M7 = %s' % fM7)
        print('   L7 = %s   (total das cinco regras)' % fL7)
        print('   G7 = %s' % fG7[:70])
        ck('M7 depende de B7 (n) e L7 (total de violacoes)',
           '$B7' in fM7 and '$L7' in fM7)
        ck('M8 depende de B8 e L8', '$B8' in fM8 and '$L8' in fM8)
        for proibido in ('Sigma', 'DPM', 'Rendimento', 'RunSize', 'Margem',
                         'Estatística!$L', 'Estatística!$V', 'Estatística!$M'):
            ck('M7/M8 NAO citam %s' % proibido,
               proibido not in fM7 and proibido not in fM8)
        ck('L7 soma as cinco colunas de regra', 'SUM(G7:K7)' in fL7, fL7)
        ck('G7 vem do Calc (motor de Westgard)', 'Calc!' in fG7)

        # ---- cenarios 1 a 4: f(B, L) responde por nivel -------------------
        print()
        print('=== CENARIOS 1 a 4: M7 e M8 respondem, e por nivel separado ===')
        print('   (B e L recebem valor controlado; a formula sob teste e M)')
        guardaB7 = tenta(lambda: pa.Range('B7').Formula)
        guardaB8 = tenta(lambda: pa.Range('B8').Formula)
        guardaG = {}
        for lin in (7, 8):
            for c in range(7, 12):
                guardaG[(lin, c)] = tenta(
                    lambda x=lin, y=c: pa.Cells(x, y).Formula)

        def cenario(n1, viol1, n2, viol2):
            tenta(lambda: pa.Range('B7').__setattr__('Value', n1))
            tenta(lambda: pa.Range('B8').__setattr__('Value', n2))
            # zera as cinco regras e poe a violacao na primeira coluna
            for lin, v in ((7, viol1), (8, viol2)):
                for c in range(7, 12):
                    tenta(lambda x=lin, y=c: pa.Cells(x, y).__setattr__('Value', 0))
                tenta(lambda x=lin, vv=v: pa.Cells(x, 7).__setattr__('Value', vv))
            tenta(lambda: xl.CalculateFull())
            return (txt(tenta(lambda: pa.Range('M7').Text)),
                    txt(tenta(lambda: pa.Range('M8').Text)))

        casos = [
            ('1. N1 com dado e SEM violacao', 50, 0, 50, 0,
             'Sem violação', 'Sem violação'),
            ('3. violacao real SO no N1', 50, 3, 50, 0,
             'REPROVA — 3 violação(ões)', 'Sem violação'),
            ('4. violacao real SO no N2', 50, 0, 50, 7,
             'Sem violação', 'REPROVA — 7 violação(ões)'),
            ('2+. os dois niveis com violacao', 50, 2, 50, 9,
             'REPROVA — 2 violação(ões)', 'REPROVA — 9 violação(ões)'),
            ('periodo sem dado nenhum', 0, 0, 0, 0,
             'sem dados no período', 'sem dados no período'),
        ]
        for rot, n1, v1, n2, v2, e7, e8 in casos:
            g7, g8 = cenario(n1, v1, n2, v2)
            ck('%s -> M7=%r' % (rot, e7), g7 == e7, repr(g7))
            ck('%s -> M8=%r' % (rot, e8), g8 == e8, repr(g8))

        print()
        print('=== os dois niveis sao independentes? ===')
        a7, a8 = cenario(50, 5, 50, 0)
        b7, b8 = cenario(50, 5, 50, 12)
        ck('mudar so o N2 nao mexeu no M7', a7 == b7, '%r / %r' % (a7, b7))
        ck('e mudou o M8', a8 != b8, '%r -> %r' % (a8, b8))

        # devolve as formulas originais
        tenta(lambda: pa.Range('B7').__setattr__('Formula', guardaB7))
        tenta(lambda: pa.Range('B8').__setattr__('Formula', guardaB8))
        for (lin, c), f in guardaG.items():
            tenta(lambda x=lin, y=c, ff=f:
                  pa.Cells(x, y).__setattr__('Formula', ff))
        tenta(lambda: xl.CalculateFull())
        ck('as formulas originais de B e G foram devolvidas',
           txt(tenta(lambda: pa.Range('B7').Formula)) == txt(guardaB7))

        # ---- cenarios 5 e 6: Sigma e DPM nao mexem no status -------------
        print()
        print('=== CENARIOS 5 e 6: Sigma baixo e DPM alto NAO reprovam ===')
        alvo = txt(tenta(lambda: es.Cells(LIN, 1).Value))
        ultEQ = tenta(lambda: cap.Cells(cap.Rows.Count, 4).End(-4162).Row)
        ultMapa = tenta(lambda: base.Cells(base.Rows.Count, 23).End(-4162).Row)
        for cc, vv in ((23, 'CAP'), (24, alvo), (25, alvo)):
            tenta(lambda c=cc, v=vv:
                  base.Cells(ultMapa + 1, c).__setattr__('Value', v))
        lin = ultEQ + 1
        for c, v in ((1, 'CAP'), (2, ROD), (3, ANO_TESTE), (4, alvo),
                     (5, '01'), (6, 130.0), (7, 100.0), (8, 5.0), (9, 6.0),
                     (10, 50.0), (11, 150.0), (12, 'Acceptable')):
            tenta(lambda cc=c, vv=v: cap.Cells(lin, cc).__setattr__('Value', vv))
        for c, f in ((15, '=IFERROR((F{0}-G{0})/G{0}*100,"")'),
                     (16, '=IF(O{0}="","",ABS(O{0}))')):
            tenta(lambda cc=c, ff=f: cap.Cells(lin, cc)
                  .__setattr__('Formula', ff.format(lin)))
        tenta(lambda: es.Range('L4').__setattr__('Value', 'CAP'))
        tenta(lambda: es.Range('N4').__setattr__('Value', ANO_TESTE))
        tenta(lambda: es.Range('P4').__setattr__('Value', ROD))
        tenta(lambda: es.Cells(LIN, 6).__setattr__('Value', 10.0))
        tenta(lambda: an.Cells(4, 20).__setattr__('Value', 12.0))
        tenta(lambda: xl.Run('AtualizarEQABase'))
        tenta(lambda: xl.CalculateFull())

        # zera as violacoes SEM mexer no n, para isolar o efeito do Sigma
        for lin2 in (7, 8):
            for c in range(7, 12):
                tenta(lambda x=lin2, y=c: pa.Cells(x, y).__setattr__('Value', 0))
        tenta(lambda: xl.CalculateFull())

        sg = tenta(lambda: es.Cells(LIN, 12).Value)
        dpm = tenta(lambda: es.Cells(LIN, 22).Value)
        m7 = txt(tenta(lambda: pa.Range('M7').Text))
        m8 = txt(tenta(lambda: pa.Range('M8').Text))
        marg = txt(tenta(lambda: es.Cells(LIN, 16).Value))
        print('   Sigma=%.4f  DPM=%.0f  margem=%r' % (sg, dpm, marg))
        print('   M7=%r   M8=%r' % (m7, m8))
        ck('o Sigma esta abaixo de 3', sg < 3, '%.4f' % sg)
        ck('o DPM esta alto (> 100.000)', dpm > 100000, '%.0f' % dpm)
        ck('cenario 5: Sigma < 3 e M7 continua "Sem violação"',
           m7 == 'Sem violação', repr(m7))
        ck('cenario 5: Sigma < 3 e M8 continua "Sem violação"',
           m8 == 'Sem violação', repr(m8))
        ck('cenario 6: DPM alto e M7/M8 nao mudaram',
           m7 == 'Sem violação' and m8 == 'Sem violação')
        ck('a margem de ETp tambem nao reprovou',
           m7 == 'Sem violação', '%s com margem %r' % (m7, marg))

        for (l2, c), f in guardaG.items():
            tenta(lambda x=l2, y=c, ff=f:
                  pa.Cells(x, y).__setattr__('Formula', ff))
        tenta(lambda: cap.Range('A%d:R%d' % (lin, lin)).ClearContents())
        tenta(lambda: xl.Run('AtualizarEQABase'))

        # ---- QA dos filtros, com valores reais ---------------------------
        print()
        print('=== QA DOS FILTROS: 3 analitos x 4 recortes ===')
        ref = openpyxl.load_workbook(referencia, data_only=True)['CAP_Evaluation_Data']
        tenta(lambda: es.Range('L4').__setattr__('Value', 'CAP'))
        tenta(lambda: es.Range('N4').__setattr__('Value', 2025))

        alvos = [('Glucose, serum', 'Glicose'),
                 ('Cholesterol', 'Colesterol total'),
                 ('Urea Nitrogen', 'Ureia')]
        for nomeRef, canon in alvos:
            porRodada = {}
            for i in range(2, ref.max_row + 1):
                if txt(ref.cell(i, 2).value) != nomeRef:
                    continue
                rd = txt(ref.cell(i, 1).value)
                r0, m0 = ref.cell(i, 4).value, ref.cell(i, 5).value
                if isinstance(r0, (int, float)) and isinstance(m0, (int, float)) and m0:
                    porRodada.setdefault(rd, []).append(
                        (ref.cell(i, 3).value, (r0 - m0) / m0 * 100))
            print()
            print('   --- %s (%s) ---' % (canon, nomeRef))
            medias = {}
            for rd in sorted(porRodada):
                itens = porRodada[rd]
                mabs = sum(abs(b) for _a, b in itens) / len(itens)
                medias[rd] = mabs
                print('      %s: %s' % (rd, '  '.join(
                    '%s bias=%+.3f |b|=%.3f' % (a, b, abs(b))
                    for a, b in itens)))
                print('         |Bias| medio da rodada = %.6f  (n=%d)'
                      % (mabs, len(itens)))
            todas = sum(medias.values()) / len(medias)
            print('      |Bias| TODAS = media das %d rodadas = %.6f'
                  % (len(medias), todas))

            # cada rodada, e depois TODAS
            for rd in sorted(medias) + ['TODAS']:
                tenta(lambda r=rd: es.Range('P4').__setattr__('Value', r))
                tenta(lambda: xl.CalculateFull())
                linha = None
                for r in range(14, 94):
                    if txt(tenta(lambda x=r: es.Cells(x, 1).Value)) == canon:
                        linha = r
                        break
                gEst = tenta(lambda x=linha: es.Cells(x, 7).Value)
                nRes = tenta(lambda x=linha: es.Cells(x, 29).Value)
                nRod = tenta(lambda x=linha: es.Cells(x, 30).Value)
                esperado = todas if rd == 'TODAS' else medias[rd]
                espRes = (sum(len(v) for v in porRodada.values())
                          if rd == 'TODAS' else len(porRodada[rd]))
                espRod = len(medias) if rd == 'TODAS' else 1
                ck('%s | %-9s |Bias|=%.6f  N res=%s  N rod=%s'
                   % (canon[:16], rd, esperado, espRes, espRod),
                   perto(gEst, esperado) and perto(nRes, espRes)
                   and perto(nRod, espRod),
                   'planilha: %s / %s / %s' % (gEst, nRes, nRod))

        # ---- propagacao Estatistica -> Painel -----------------------------
        print()
        print('=== QA DA PROPAGACAO ESTATISTICA -> PAINEL ===')
        tenta(lambda: es.Range('P4').__setattr__('Value', 'TODAS'))
        for _nomeRef, canon in alvos:
            idx = None
            for k in range(4, 44):
                if txt(tenta(lambda x=k: an.Cells(x, 1).Value)) == canon:
                    idx = k - 3
                    break
            if idx is None:
                continue
            tenta(lambda i=idx: pa.Range('B3').__setattr__('Value', i))
            tenta(lambda: xl.CalculateFull())
            linha = None
            for r in range(14, 94):
                if txt(tenta(lambda x=r: es.Cells(x, 1).Value)) == canon:
                    linha = r
                    break
            campos = [('CV %', 6, 10, 19), ('|Bias| %', 7, 10, 20),
                      ('ETp %', 11, 10, 21), ('Sigma', 12, 10, 22),
                      ('Classificação', 13, 10, 23), ('DPM', 22, 10, 24),
                      ('ET %', 8, 17, 19), ('Margem p.p.', 14, 17, 21),
                      ('Margem %', 15, 17, 22), ('Situação', 16, 17, 23),
                      ('Regras', 24, 22, 20), ('N', 25, 22, 21),
                      ('Run size', 26, 22, 22), ('Cobertura', 27, 22, 24)]
            iguais = 0
            for rot, cEst, rPai, cPai in campos:
                a = tenta(lambda x=linha, y=cEst: es.Cells(x, y).Value)
                b = tenta(lambda x=rPai, y=cPai: pa.Cells(x, y).Value)
                ok = perto(a, b) if isinstance(a, (int, float)) else txt(a) == txt(b)
                if ok:
                    iguais += 1
                else:
                    ck('%s / %s: Estatistica == Painel' % (canon[:14], rot),
                       False, '%r vs %r' % (a, b))
            ck('%s: %d de %d campos identicos entre Estatistica e Painel'
               % (canon[:16], iguais, len(campos)), iguais == len(campos))
    finally:
        try:
            wb.Close(False)
        except Exception:
            pass
        try:
            xl.Quit()
        except Exception:
            pass

    print()
    print('=' * 70)
    if falhas:
        print('FALHAS (%d):' % len(falhas))
        for f in falhas:
            print('   - %s' % f)
        sys.exit(1)
    print('STATUS WESTGARD E FILTROS DE EQA: TODAS AS PROVAS PASSARAM')


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
