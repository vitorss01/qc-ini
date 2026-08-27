# -*- coding: utf-8 -*-
"""inspecionar_fase1.py - o estado REAL dos achados pendentes da Fase 1

Nao corrige nada. So mede, nos TRES produtos, o que a FASE1_documentacao.md
deixou registrado como pendente ou verificado:

  #4  Calc usa referencia absoluta fixa ($BT$1/$BU$1) para media e DP
  #5  LiberStore gravado so na troca de lote / BeforeSave
  #6  modulos de classe de planilha sem codigo
  #8  coluna RUN com formato de data (corrigido -- confere se continua)
  SSOT  aba Resultados com o schema A=RUN .. H=NC
  camadas  mDados / mUI / mLotes / mSeguranca presentes

Le com openpyxl e oletools: nao abre o Excel, entao nao altera nada.
"""
import io
import os
import re
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', write_through=True)
import openpyxl
from oletools.olevba import VBA_Parser

PRODUTOS = ['QC_Hematologia.xlsm', 'QC_Bioquimica.xlsm', 'QC_Imunologia.xlsm']

SCHEMA = ['RUN', 'Data', 'Nível', 'Lote', 'Analito', 'Resultado', 'Status', 'NC']
CAMADAS = ['mDados', 'mUI', 'mLotes', 'mSeguranca']


def vba(arq):
    """{nome_do_modulo: codigo}"""
    p = VBA_Parser(arq)
    d = {}
    for _, _, nome, cod in p.extract_macros():
        if nome:
            d[nome.replace('.bas', '').replace('.cls', '').replace('.frm', '')] = cod or ''
    p.close()
    return d


def secao(t):
    print()
    print('=' * 78)
    print(t)
    print('=' * 78)


def main():
    base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    achados = []

    for arq in PRODUTOS:
        cam = os.path.join(base, arq)
        if not os.path.exists(cam):
            print('AUSENTE: %s' % arq)
            continue
        secao(arq)
        wb = openpyxl.load_workbook(cam)
        mods = vba(cam)
        prod = arq.replace('QC_', '').replace('.xlsm', '')

        # ---- SSOT: a aba Resultados e o schema da Fase 1 -----------------
        if 'Resultados' not in wb.sheetnames:
            achados.append((prod, 'SSOT', 'aba Resultados ausente'))
            print('  SSOT: aba Resultados AUSENTE')
        else:
            ws = wb['Resultados']
            cab = [str(ws.cell(1, c).value or '').strip() for c in range(1, 9)]
            bate = [a.lower() == b.lower() for a, b in zip(cab, SCHEMA)]
            print('  SSOT Resultados A1:H1 = %s' % cab)
            if not all(bate):
                achados.append((prod, 'SSOT', 'schema divergente: %s' % cab))
                print('       DIVERGE do schema da Fase 1 %s' % SCHEMA)
            else:
                print('       schema da Fase 1 conferido')

            # #8 -- RUN nao pode ter formato de data
            fmts = {ws.cell(r, 1).number_format for r in range(2, min(60, ws.max_row + 1))}
            ruins = [f for f in fmts if any(x in str(f).lower() for x in ('d', 'y', 'm/'))
                     and str(f) != 'General']
            print('  #8 formato da coluna RUN: %s' % sorted(fmts))
            if ruins:
                achados.append((prod, '#8', 'RUN com formato de data: %s' % ruins))
                print('       FORMATO DE DATA DE VOLTA: %s' % ruins)

            # exclusao logica viva?
            col_st = [str(ws.cell(r, 7).value or '') for r in range(2, ws.max_row + 1)]
            vals = sorted({v for v in col_st if v})
            n_ativo = sum(1 for v in col_st if v.lower().startswith('ativ'))
            n_exc = sum(1 for v in col_st if v.lower().startswith('exclu'))
            print('  Status: %s  (ativos %d, excluidos %d, linhas %d)'
                  % (vals[:4], n_ativo, n_exc, len([v for v in col_st if v])))

        # ---- camadas do VBA ---------------------------------------------
        falta = [m for m in CAMADAS if m not in mods]
        print('  camadas da Fase 1: %s'
              % ('todas presentes' if not falta else 'FALTAM %s' % falta))
        if falta:
            achados.append((prod, 'camadas', 'modulos ausentes: %s' % falta))
        if 'mSystem' in mods:
            achados.append((prod, 'camadas', 'mSystem (monolito) ainda presente'))
            print('       mSystem AINDA PRESENTE (o monolito deveria ter saido)')

        # ---- #6: classes de planilha sem codigo -------------------------
        vazias, comcod = [], []
        for nome, cod in mods.items():
            if not re.match(r'^(Planilha|Sheet|Plan)\d+$', nome) and nome != 'EstaPastaDeTrabalho':
                continue
            corpo = [l for l in cod.split('\n')
                     if l.strip() and not l.strip().startswith(("'", 'Attribute',
                                                                'VERSION', 'BEGIN',
                                                                'END', 'Option'))]
            (comcod if corpo else vazias).append(nome)
        print('  #6 classes de planilha: %d sem codigo, %d com codigo'
              % (len(vazias), len(comcod)))
        if vazias:
            achados.append((prod, '#6', '%d classes de planilha sem codigo' % len(vazias)))

        # ---- #4: referencia absoluta fixa no Calc ------------------------
        if 'Calc' in wb.sheetnames:
            ws = wb['Calc']
            achou = {}
            lim = min(ws.max_row, 200)
            for linha in ws.iter_rows(min_row=1, max_row=lim):
                for c in linha:
                    v = c.value
                    if isinstance(v, str) and v.startswith('='):
                        for m in re.finditer(r'\$B[TU]\$1\b', v):
                            achou.setdefault(m.group(0), []).append(c.coordinate)
            if achou:
                total = sum(len(v) for v in achou.values())
                achados.append((prod, '#4', '%d celula(s) com $BT$1/$BU$1' % total))
                print('  #4 referencia absoluta fixa: %d celula(s) %s'
                      % (total, {k: v[:3] for k, v in achou.items()}))
            else:
                print('  #4 nenhuma referencia $BT$1/$BU$1 no Calc (primeiras %d linhas)' % lim)
        else:
            print('  #4 aba Calc ausente')

        # ---- #5: LiberStore so no BeforeSave -----------------------------
        alvo = [n for n in mods if 'lote' in n.lower() or 'estapasta' in n.lower()]
        chamadas = {}
        for nome, cod in mods.items():
            for m in re.finditer(r'\b(SalvarViewNoBloco|FlushLoteAtual|FlushLiber\w*)\b', cod or ''):
                chamadas.setdefault(m.group(1), set()).add(nome)
        print('  #5 quem chama o flush de views: %s'
              % {k: sorted(v) for k, v in chamadas.items()})

        wb.close()

    secao('RESUMO DOS ACHADOS ABERTOS')
    if not achados:
        print('  nenhum achado aberto')
    for prod, cod, txt in achados:
        print('  [%-11s] %-8s %s' % (prod, cod, txt))
    print()
    print('total: %d achado(s)' % len(achados))


main()
