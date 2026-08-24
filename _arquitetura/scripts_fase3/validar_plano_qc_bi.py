# -*- coding: utf-8 -*-
"""validar_plano_qc_bi.py - o plano de CQ publicado no BI esta correto?

O QUE ESTE SCRIPT MEDE

Le tblBI_Fato direto do artefato e confere, por analito e lote, se o plano de
CQ publicado corresponde ao Sigma do PIOR nivel (ADR-040). Nao le a planilha
pela interface nem o modelo do Power BI: mede o CONTRATO, que e o que o BI
consome.

BLOCOS (correspondem ao roteiro de validacao)

  A. Fronteiras Sigma ..... 2,99 / 3,00 / 3,01 ... 6,00 / 6,01, contra a
                            matriz Cfg_PlanoQC lida do proprio arquivo.
  B. Pior nivel ........... Sigma_Plano == MIN(Sigma dos niveis validos) e
                            Nivel_Governante aponta o nivel desse minimo.
  C. Ausentes ............. sem Sigma valido em nenhum nivel -> SEM DADOS,
                            plano vazio, cinco flags FALSE.
  D. Analitos reais ....... Lactato e companhia, com os dois Sigmas visiveis.
  E. Sigma x regras ....... a cadeia Regra_Westgard_Recomendada bate com a
                            faixa de Sigma_Plano.
  F. Flags booleanas ...... Usar_* concorda com a cadeia de regras.
  G. N / Run Size ......... coerentes com a faixa; VAZIOS abaixo de 3 Sigma.

Uso:
    python validar_plano_qc_bi.py
    python validar_plano_qc_bi.py --produto Bioquimica
"""
import argparse
import io
import os
import sys
import warnings

warnings.filterwarnings('ignore')
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8',
                              write_through=True)

PERFIL = os.environ.get('USERPROFILE', os.path.expanduser('~'))
ARTEFATOS = (
    ('Bioquimica', os.path.join(PERFIL, 'QCINI_build_hardening1_Bioquimica',
                                'QC_Bioquimica.xlsm')),
    ('Hematologia', os.path.join(PERFIL, 'QCINI_build_hardening1_Hematologia',
                                 'QC_Hematologia.xlsm')),
)

# Sigmas de fronteira exigidos pelo roteiro. Existem para pegar erro de
# > contra >= : 3,00 pertence a faixa "Marginal" e 2,99 nao.
FRONTEIRAS = [2.99, 3.00, 3.01, 3.99, 4.00, 4.01,
              4.99, 5.00, 5.01, 5.99, 6.00, 6.01, 6.50]

REGRAS = ('1_3s', '2_2s', 'R_4s', '4_1s', '8x')

resultados = []


def anota(bloco, teste, entrada, esperado, obtido):
    ok = (str(esperado).strip() == str(obtido).strip())
    resultados.append((bloco, teste, entrada, esperado, obtido, ok))
    return ok


def norm(t):
    """Normaliza a cadeia de regras: separador e traco nao devem decidir
    igualdade. '4-1s' e '4_1s' sao a mesma regra."""
    s = str(t or '').lower().replace('/', ' ').replace(';', ' ')
    s = s.replace(',', ' ').replace('-', '_')
    return ' '.join(s.split())


def faixa_da_matriz(matriz, sigma):
    """A linha do plano para este Sigma, pela mesma comparacao do motor:
    Sigma_Min <= s < Sigma_Max."""
    if sigma is None:
        return None
    for lin in matriz:
        smin, smax = lin['min'], lin['max']
        if smin is None or smax is None:
            continue
        if smin <= sigma < smax:
            return lin
    return None


def ler_matriz(wb):
    """A matriz do plano so existe onde o ADR-035 foi aplicado. Ausente, o
    produto nao tem plano de CQ a publicar -- e isso e um achado a relatar,
    nao uma excecao a estourar no meio do teste."""
    if 'Cfg_PlanoQC' not in wb.sheetnames:
        return None
    ws = wb['Cfg_PlanoQC']
    matriz = []
    for linha in ws.iter_rows(min_row=4, max_row=12, max_col=6, values_only=True):
        smin, smax = linha[0], linha[1]
        if not isinstance(smin, (int, float)) or not isinstance(smax, (int, float)):
            continue
        matriz.append({'min': float(smin), 'max': float(smax),
                       'classe': linha[2], 'regras': linha[3],
                       'n': linha[4], 'run': linha[5]})
    return matriz


def ler_fato(wb):
    """iter_rows e nao ws.cell(): em modo read_only o acesso por coordenada
    percorre a linha inteira a cada chamada, e 6.808 x 84 celulas assim viram
    tempo quadratico -- a primeira versao deste script nao terminava."""
    ws = wb['BI_Data']
    it = ws.iter_rows(values_only=True)
    try:
        cabecalho = next(it)
    except StopIteration:
        return {}, []
    cab = {}
    for i, v in enumerate(cabecalho):
        if v is not None and str(v).strip():
            cab[str(v)] = i + 1
    idx = [(k, c - 1) for k, c in cab.items()]
    linhas = []
    for linha in it:
        if not linha or linha[0] is None:
            continue
        linhas.append({k: linha[i] for k, i in idx})
    return cab, linhas


def num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def vazio(v):
    return v is None or str(v).strip() == ''


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--produto', default=None)
    a = ap.parse_args()

    import openpyxl

    for produto, caminho in ARTEFATOS:
        if a.produto and produto != a.produto:
            continue
        print('=' * 70)
        print(produto, ' ', os.path.basename(caminho))
        print('=' * 70)
        if not os.path.isfile(caminho):
            print('   artefato ausente -- BLOQUEADO')
            continue
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
        try:
            matriz = ler_matriz(wb)
            cab, linhas = ler_fato(wb)
        finally:
            wb.close()

        obrig = ['Sigma', 'Sigma_Plano', 'Nivel_Governante',
                 'Classificacao_Sigma_Plano', 'Regra_Westgard_Recomendada',
                 'N_Controle_Recomendado', 'RunSize_Max_Recomendado'] + \
                ['Usar_' + x for x in REGRAS]
        falta = [c for c in obrig if c not in cab]
        if falta:
            print('   contrato incompleto, faltam: %s' % ', '.join(falta))
            print('   -> BLOQUEADO (rebuild com ADR-040 pendente)')
            continue
        print('   %d linhas, %d colunas' % (len(linhas), len(cab)))

        # ---- A. fronteiras, contra a matriz do proprio arquivo
        print('\n   A. FRONTEIRAS SIGMA')
        if matriz is None:
            print('      nao se aplica: este produto nao tem Cfg_PlanoQC')
        for s in (FRONTEIRAS if matriz else []):
            lin = faixa_da_matriz(matriz, s)
            classe = lin['classe'] if lin else '(sem faixa)'
            regras = norm(lin['regras']) if lin else ''
            n_esp = '' if not lin or vazio(lin['n']) else str(lin['n'])
            r_esp = '' if not lin or vazio(lin['run']) else str(lin['run'])
            print('      %-5s -> %-22s | %-32s | N=%-4s Run=%-5s'
                  % (s, classe, regras, n_esp or '(vazio)', r_esp or '(vazio)'))
            if s < 3:
                anota('A', 'faixa <3 declara as cinco regras', s,
                      norm('1_3s / 2_2s / R_4s / 4_1s / 8x'), regras)
                anota('A', 'faixa <3 sem N', s, '', n_esp)
                anota('A', 'faixa <3 sem run size', s, '', r_esp)

        # ---- B/C/D: por grupo (analito, lote)
        print('\n   B. PIOR NIVEL GOVERNA')
        grupos = {}
        for L in linhas:
            ch = (L.get('ID_Analito'), L.get('ID_Lote'))
            grupos.setdefault(ch, []).append(L)

        piores, sem_dados = 0, 0
        for ch, rows in sorted(grupos.items(), key=lambda x: str(x[0])):
            sigmas = {}
            for L in rows:
                s = num(L.get('Sigma'))
                if s is not None and s > 0:
                    sigmas[L.get('Nivel')] = s
            sp = num(rows[0].get('Sigma_Plano'))
            gov = rows[0].get('Nivel_Governante')

            if sigmas:
                esp_sp = min(sigmas.values())
                esp_nv = 'Nivel %s' % min(sigmas, key=lambda k: sigmas[k])
                ok1 = anota('B', 'Sigma_Plano = MIN dos niveis', str(ch),
                            round(esp_sp, 6), round(sp, 6) if sp is not None else None)
                ok2 = anota('B', 'Nivel_Governante', str(ch), esp_nv, gov)
                if not (ok1 and ok2):
                    print('      FALHA %s: sigmas=%s  plano=%s  gov=%s'
                          % (ch[0], sigmas, sp, gov))
                piores += 1
            else:
                anota('C', 'sem Sigma -> SEM DADOS', str(ch), 'SEM DADOS', gov)
                anota('C', 'sem Sigma -> plano vazio', str(ch), True,
                      vazio(rows[0].get('Regra_Westgard_Recomendada')))
                for rg in REGRAS:
                    anota('C', 'sem Sigma -> Usar_%s falso' % rg, str(ch),
                          False, bool(rows[0].get('Usar_' + rg)))
                sem_dados += 1

            # ---- E/F/G: plano coerente com a faixa de Sigma_Plano
            if sp is not None and matriz:
                lin = faixa_da_matriz(matriz, sp)
                if lin:
                    anota('E', 'regras batem com a faixa', str(ch),
                          norm(lin['regras']),
                          norm(rows[0].get('Regra_Westgard_Recomendada')))
                    anota('G', 'N bate com a faixa', str(ch),
                          '' if vazio(lin['n']) else str(lin['n']),
                          '' if vazio(rows[0].get('N_Controle_Recomendado'))
                          else str(rows[0].get('N_Controle_Recomendado')))
                    anota('G', 'run size bate com a faixa', str(ch),
                          '' if vazio(lin['run']) else str(lin['run']),
                          '' if vazio(rows[0].get('RunSize_Max_Recomendado'))
                          else str(rows[0].get('RunSize_Max_Recomendado')))
                    toks = norm(lin['regras']).split()
                    for rg in REGRAS:
                        anota('F', 'Usar_%s concorda com a cadeia' % rg, str(ch),
                              norm(rg) in toks, bool(rows[0].get('Usar_' + rg)))
        print('      %d grupos com Sigma, %d sem dado' % (piores, sem_dados))

        # ---- D. analitos reais do roteiro
        print('\n   D. ANALITOS REAIS')
        alvo = ['Lactato', 'Ácido úrico', 'Cálcio', 'Bilirrubina total',
                'Capacidade de fixação do ferro']
        for nome in alvo:
            rows = [L for L in linhas if str(L.get('Analito')) == nome]
            if not rows:
                continue
            sig = {}
            for L in rows:
                s = num(L.get('Sigma'))
                if s is not None and s > 0:
                    sig[L.get('Nivel')] = round(s, 3)
            r0 = rows[0]
            print('      %-32s N1=%-8s N2=%-8s | plano=%-8s %-9s | %s'
                  % (nome[:32],
                     sig.get(1, '--'), sig.get(2, '--'),
                     round(num(r0.get('Sigma_Plano')), 3)
                     if num(r0.get('Sigma_Plano')) is not None else '--',
                     r0.get('Nivel_Governante'),
                     norm(r0.get('Regra_Westgard_Recomendada')) or '(vazio)'))
            print('%s N=%s  Run=%s  %s' % (' ' * 38,
                  r0.get('N_Controle_Recomendado') or '(vazio)',
                  r0.get('RunSize_Max_Recomendado') or '(vazio)',
                  r0.get('Classificacao_Sigma_Plano')))

    # ---------------------------------------------------------- resumo
    print()
    print('=' * 70)
    print('MATRIZ DE VALIDACAO')
    print('=' * 70)
    blocos = {}
    for b, t, e, esp, obt, ok in resultados:
        d = blocos.setdefault(b, [0, 0])
        d[0 if ok else 1] += 1
    for b in sorted(blocos):
        p, f = blocos[b]
        print('   bloco %s: %d PASS, %d FAIL' % (b, p, f))
    falhas = [r for r in resultados if not r[5]]
    if falhas:
        print('\n   PRIMEIRAS FALHAS:')
        for b, t, e, esp, obt, ok in falhas[:12]:
            print('      [%s] %s | entrada=%s | esperado=%s | obtido=%s'
                  % (b, t, e, esp, obt))
    print()
    print('TOTAL: %d PASS, %d FAIL' % (len(resultados) - len(falhas), len(falhas)))
    return 1 if falhas else 0


if __name__ == '__main__':
    sys.exit(main())
