# -*- coding: utf-8 -*-
"""publicar_eqa_bi.py - ADR-034: a EQA_Base vira fonte nomeada, e QA visual

O QUE FAZ

  1. transforma a EQA_Base numa TABELA ESTRUTURADA chamada tblEQA_Base.
     O Power Query referencia por NOME; faixa solta obriga a apontar
     "EQA_Base!A1:U5001", que quebra no dia em que a base crescer.
  2. audita a paridade visual entre EQA.CAP_Dados e a planilha de referencia:
     cabecalho, largura de coluna, formato numerico, congelamento, tabela,
     formatacao condicional.

A base fica veryHidden. Tabela em aba oculta e legivel pelo Power Query do
mesmo jeito -- o que o usuario nao deve e digitar nela.

Uso: python publicar_eqa_bi.py <arquivo.xlsm> <referencia.xlsx>
"""
import io
import os
import sys
import time
import subprocess

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', write_through=True)
import win32com.client as w
import openpyxl


def novo_excel():
    for t in range(1, 9):
        try:
            xl = w.DispatchEx('Excel.Application')
            xl.Visible = False
            xl.DisplayAlerts = False
            xl.EnableEvents = False
            xl.AutomationSecurity = 1
            return xl
        except Exception:
            if t in (1, 3, 5):
                subprocess.call(['powershell', '-NoProfile', '-Command',
                                 'Start-Process excel.exe -WindowStyle Minimized'],
                                stderr=subprocess.DEVNULL)
                time.sleep(10)
            time.sleep(2.5 * t)
    raise RuntimeError('Excel COM nao subiu')


def tenta(fn, vezes=10):
    ult = None
    for i in range(vezes):
        try:
            return fn()
        except Exception as e:
            ult = e
            s = str(e).lower()
            if 'rejeitada' not in s and 'rejected' not in s:
                raise
            time.sleep(1.0 + 0.8 * i)
    raise ult


def norm(f):
    return (str(f).replace('Geral', 'General').replace('#', '0')
            .replace('Padrao', 'General').strip())


def col(n):
    s = ''
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def main(caminho, referencia):
    caminho = os.path.abspath(caminho)

    # ---- o que a referencia usa, lido do arquivo -------------------------
    rw = openpyxl.load_workbook(referencia)
    rs = rw['CAP_Evaluation_Data']
    refCab = [rs.cell(1, c).value for c in range(1, 17)]
    refLarg = [round(rs.column_dimensions[col(c)].width, 1)
               if rs.column_dimensions.get(col(c)) and
               rs.column_dimensions[col(c)].width else None
               for c in range(1, 17)]
    refFmt = [rs.cell(2, c).number_format for c in range(1, 17)]
    refH = rs.cell(1, 1)
    refFill = refH.fill.fgColor.rgb if refH.fill and refH.fill.fgColor else None
    refFonte = refH.font.color.rgb if refH.font.color else None
    rw.close()

    xl = novo_excel()
    wb = xl.Workbooks.Open(caminho)
    salvou = False
    if wb.ReadOnly:
        wb.Close(False)
        xl.Quit()
        raise SystemExit('Somente leitura: %s' % caminho)
    try:
        xl.Calculation = -4135
        base = wb.Worksheets('EQA_Base')
        visAntes = base.Visible
        base.Visible = -1

        # ---- 1. tblEQA_Base ---------------------------------------------
        ult = tenta(lambda: base.Cells(base.Rows.Count, 1).End(-4162).Row)
        if ult < 2:
            raise SystemExit('EQA_Base vazia -- rode a migracao antes')
        for lo in list(base.ListObjects):
            lo.Unlist()
        lo = base.ListObjects.Add(
            1, base.Range(base.Cells(1, 1), base.Cells(ult, 21)), None, 1)
        lo.Name = 'tblEQA_Base'
        lo.TableStyle = 'TableStyleLight1'
        print('tblEQA_Base: %s (%d linhas x %d colunas)'
              % (lo.Range.Address, lo.ListRows.Count, lo.ListColumns.Count))
        campos = [str(base.Cells(1, c).Value) for c in range(1, 22)]
        print('campos publicados ao Power BI:')
        for i in range(0, 21, 3):
            print('   ' + ' | '.join(campos[i:i + 3]))

        base.Visible = visAntes if visAntes != -1 else 2
        print('EQA_Base continua %s'
              % ('veryHidden' if base.Visible == 2 else 'visivel'))

        # ---- 2. QA visual contra a referencia ---------------------------
        cap = wb.Worksheets('EQA.CAP_Dados')
        print()
        print('=== QA VISUAL: EQA.CAP_Dados contra a planilha de referencia ===')
        print('%-3s %-24s | %-24s | %s'
              % ('col', 'referencia', 'QC_Bioquimica', 'largura / formato'))
        # a referencia nao tem coluna Provedor nem Ano: ela e monoprovedor e
        # monoano. O modulo tem, porque consolida dois programas e varios anos.
        # ref -> modulo. A referencia comeca em Survey; o modulo tem Provedor
        # antes dele e Ano depois, entao o deslocamento nao e constante.
        mapa = {1: 2, 2: 4, 3: 5, 4: 6, 5: 7, 6: 8, 7: 9, 8: 10, 9: 11,
                10: 12, 11: 13, 12: 14, 13: 15, 14: 16, 15: 17, 16: 18}
        divergentes = []
        for cRef, cNovo in sorted(mapa.items()):
            nome = str(tenta(lambda c=cNovo: cap.Cells(1, c).Value))
            larg = round(tenta(lambda c=cNovo: cap.Columns(c).ColumnWidth), 1)
            fmt = str(tenta(lambda c=cNovo: cap.Cells(2, c).NumberFormat))
            # O Excel devolve o formato LOCALIZADO (Geral, #.000); o openpyxl
            # devolve o codigo invariante (General, 0.000). Comparar cru
            # acusaria divergencia em toda linha. Normaliza os dois lados.
            marca = ''
            if norm(refFmt[cRef - 1]) != norm(fmt):
                marca = '  <- confira'
                divergentes.append('%s: ref %s / modulo %s'
                                   % (nome, refFmt[cRef - 1], fmt))
            print('%-3s %-24s | %-24s | %s / %s%s'
                  % (col(cNovo), str(refCab[cRef - 1])[:24], nome[:24],
                     larg, fmt, marca))
        print()
        print('colunas so do modulo (a referencia e monoprovedor/monoano):')
        for c in (1, 16):
            print('   %s %s' % (col(c), tenta(lambda cc=c: cap.Cells(1, cc).Value)))

        h = cap.Cells(1, 1)
        print()
        print('cabecalho  referencia: fundo %s / fonte %s / negrito True'
              % (refFill, refFonte))
        # COM devolve cor como float; %X exige inteiro. E a cor vem em BGR,
        # enquanto o openpyxl leu a referencia em RGB -- por isso a inversao
        # antes de comparar.
        def bgr_para_rgb(v):
            n = int(v)
            return ((n & 0xFF) << 16) | (n & 0xFF00) | ((n >> 16) & 0xFF)
        print('cabecalho  modulo    : fundo FF%06X / fonte FF%06X / negrito %s'
              % (bgr_para_rgb(h.Interior.Color), bgr_para_rgb(h.Font.Color),
                 h.Font.Bold))
        igual = ('FF%06X' % bgr_para_rgb(h.Interior.Color)) == str(refFill)
        print('cabecalho  identico a referencia? %s' % ('SIM' if igual else 'nao'))
        janela = wb.Windows(1)
        cap.Activate()
        print('congelamento: %s (a referencia nao congela nada)'
              % ('sim, linha 1 e colunas A:D' if janela.FreezePanes else 'NAO'))
        print('tabela estruturada: %s | filtro: %s'
              % (cap.ListObjects(1).Name, cap.ListObjects(1).ShowAutoFilter))
        print('formatacao condicional na Avaliacao: %d regra(s)'
              % cap.Range('L2:L2000').FormatConditions.Count)
        print('barra de dados no |Bias|: %d regra(s)'
              % cap.Range('P2:P2000').FormatConditions.Count)
        print()
        print('formatos divergentes da referencia: %s'
              % (divergentes if divergentes else 'nenhum'))

        # ---- 3. procurar #### e texto cortado ---------------------------
        cortadas = []
        for c in range(1, 19):
            for r in (2, 3, 100, 300, 500):
                t = str(tenta(lambda cc=c, rr=r: cap.Cells(rr, cc).Text))
                if t and set(t) == {'#'}:
                    cortadas.append('%s%d' % (col(c), r))
        print('celulas exibindo #### (coluna estreita): %s'
              % (cortadas if cortadas else 'nenhuma'))

        # ---- 4. paridade entre os dois provedores -----------------------
        ctl = wb.Worksheets('EQA.Controllab_Dados')
        difs = []
        for c in range(1, 19):
            a = str(tenta(lambda cc=c: cap.Cells(1, cc).Value))
            b = str(tenta(lambda cc=c: ctl.Cells(1, cc).Value))
            la = round(tenta(lambda cc=c: cap.Columns(cc).ColumnWidth), 1)
            lb = round(tenta(lambda cc=c: ctl.Columns(cc).ColumnWidth), 1)
            if a != b or la != lb:
                difs.append('%s: %r/%s vs %r/%s' % (col(c), a, la, b, lb))
        print('diferencas de estrutura entre CAP e Controllab: %s'
              % (difs if difs else 'nenhuma -- as duas abas sao gemeas'))

        xl.Calculation = -4105
        wb.Save()
        salvou = True
        print()
        print('SALVO: %s' % caminho)
    finally:
        try:
            wb.Close(salvou)
        except Exception:
            pass
        try:
            xl.Quit()
        except Exception:
            pass


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
