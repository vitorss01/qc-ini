# -*- coding: utf-8 -*-
"""sonda_fase1b.py - cabecalho real do DB_Resultados e o padrao de Z-score vigente"""
import io
import os
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', write_through=True)
import openpyxl

PRODUTOS = ['QC_Hematologia.xlsm', 'QC_Bioquimica.xlsm', 'QC_Imunologia.xlsm']
SCHEMA = ['RUN', 'Data', 'Nível', 'Lote', 'Analito', 'Resultado', 'Status', 'NC']

base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

for arq in PRODUTOS:
    cam = os.path.join(base, arq)
    if not os.path.exists(cam):
        continue
    wb = openpyxl.load_workbook(cam)
    print()
    print('=' * 78)
    print(arq)
    print('=' * 78)

    ws = wb['DB_Resultados']
    lin = 0
    for r in range(1, 8):
        v = str(ws.cell(r, 1).value or '').strip().lower()
        if v == 'run':
            lin = r
            break
    if lin:
        cab = [str(ws.cell(lin, c).value or '').strip() for c in range(1, 9)]
        ok = all(a.lower() == b.lower() for a, b in zip(cab, SCHEMA))
        print('  DB_Resultados: cabecalho na linha %d' % lin)
        print('     %s' % cab)
        print('     schema da Fase 1: %s' % ('CONFERE' if ok else 'DIVERGE de %s' % SCHEMA))
    else:
        print('  DB_Resultados: nao achei "RUN" nas 7 primeiras linhas da coluna A')
        for r in range(1, 5):
            print('     linha %d: %s'
                  % (r, [str(ws.cell(r, c).value or '')[:18] for c in range(1, 9)]))

    # a aba Resultados (view) espelha?
    if 'Resultados' in wb.sheetnames:
        v = wb['Resultados']
        print('  Resultados (view): A1=%r  A2=%r'
              % (str(v['A1'].value or '')[:50], str(v['A2'].value or '')[:50]))
        print('     A3=%r' % str(v['A3'].value or '')[:90])

    # padrao de Z-score no Calc
    if 'Calc' in wb.sheetnames:
        c = wb['Calc']
        print('  Calc G3 = %s' % str(c['G3'].value)[:160])
        print('  Calc Q3 = %s' % str(c['Q3'].value)[:160])
    wb.close()
