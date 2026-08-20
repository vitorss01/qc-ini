# -*- coding: utf-8 -*-
"""validar_bi_x_excel.py - os numeros do Power BI batem com a aba Estatistica?

POR QUE ESTA COMPARACAO E A QUE IMPORTA

As medidas DAX NAO recalculam estatistica: elas agregam colunas que o mBI ja
gravou por linha (Media_Observada, DP_Observado, CV_Observado_pct,
Bias_Observado_pct, Sigma_Obs). E a arquitetura certa -- o motor do Excel e a
fonte, o BI consome. Consequencia pratica: se as colunas da BI_Data baterem com
a aba Estatistica, o Power BI bate por construcao.

Entao a divergencia so pode nascer em um lugar: entre a aba Estatistica e a
BI_Data. E exatamente isso que este script mede, analito a analito, nivel a
nivel, nos dois produtos.

O QUE E COMPARADO

    Estatistica!C  n        <->  BI_Data.N_Observado
    Estatistica!D  media    <->  BI_Data.Media_Observada
    Estatistica!E  DP       <->  BI_Data.DP_Observado
    Estatistica!F  CV%      <->  BI_Data.CV_Observado_pct
    Estatistica!K  Bias%    <->  BI_Data.Bias_Observado_pct
    Estatistica!N  ET%      <->  BI_Data.ET_Observado_pct
    Estatistica!R  Sigma    <->  BI_Data.Sigma_Obs

Tolerancia 1e-6 no valor relativo. Diferenca maior nao e arredondamento.

Uso: python validar_bi_x_excel.py
"""
import io
import os
import sys
import collections

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', write_through=True)
import openpyxl

PERFIL = os.path.expanduser('~')
ARTEFATOS = (
    ('Bioquimica', os.path.join(PERFIL, 'QCINI_build_hardening1_Bioquimica', 'QC_Bioquimica.xlsm')),
    ('Hematologia', os.path.join(PERFIL, 'QCINI_build_hardening1_Hematologia', 'QC_Hematologia.xlsm')),
)

# coluna da aba Estatistica -> nome da coluna na BI_Data
PARES = [
    ('n',      3,  'N_Observado'),
    ('media',  4,  'Media_Observada'),
    ('DP',     5,  'DP_Observado'),
    ('CV%',    6,  'CV_Observado_pct'),
    ('Bias%', 11,  'Bias_Observado_pct'),
    ('ET%',   14,  'ET_Observado_pct'),
    ('Sigma', 18,  'Sigma_Obs'),
]
TOL = 1e-6
falhas = []


def num(x):
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    return None


def perto(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    escala = max(abs(a), abs(b), 1.0)
    return abs(a - b) / escala <= TOL


def main():
    for produto, caminho in ARTEFATOS:
        print('=' * 70)
        print('%s   %s' % (produto, os.path.basename(caminho)))
        print('=' * 70)
        if not os.path.exists(caminho):
            print('   artefato ausente -- pulado')
            falhas.append('%s: artefato ausente' % produto)
            continue
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)

        # --- BI_Data: uma linha por (analito, nivel, lote) basta -----------
        ws = wb['BI_Data']
        it = ws.iter_rows(values_only=True)
        cab = next(it)
        ix = {c: i for i, c in enumerate(cab) if c}
        bi = {}
        for r in it:
            if r[0] in (None, ''):
                continue
            chave = (str(r[ix['Analito']]).strip(), int(r[ix['Nivel']]))
            if chave in bi:
                continue
            bi[chave] = {nome: num(r[ix[nome]]) for _, _, nome in PARES if nome in ix}
        print('   BI_Data: %d grupos (analito, nivel)' % len(bi))

        # --- aba Estatistica ----------------------------------------------
        es = wb['Estatística'] if 'Estatística' in wb.sheetnames else wb['Estatistica']
        linhas = 0
        comparados = 0
        div = collections.Counter()
        exemplos = []
        for r in es.iter_rows(min_row=14, max_row=200, max_col=18, values_only=True):
            nome = r[0]
            if nome in (None, ''):
                continue
            try:
                nivel = int(r[1])
            except Exception:
                continue
            linhas += 1
            chave = (str(nome).strip(), nivel)
            if chave not in bi:
                continue
            for rot, col, campo in PARES:
                if campo not in bi[chave]:
                    continue
                a = num(r[col - 1])
                b = bi[chave][campo]
                if a is None and b is None:
                    continue
                comparados += 1
                if not perto(a, b):
                    div[rot] += 1
                    if len(exemplos) < 6:
                        exemplos.append('%s N%d %s: excel=%r  bi=%r'
                                        % (nome, nivel, rot, a, b))
        print('   Estatistica: %d linhas ; %d valores comparados' % (linhas, comparados))
        if div:
            print('   DIVERGENCIAS: %s' % dict(div))
            for e in exemplos:
                print('      %s' % e)
            falhas.append('%s: %d divergencia(s)' % (produto, sum(div.values())))
        else:
            print('   divergencias: NENHUMA')
        if comparados == 0:
            falhas.append('%s: nada foi comparado -- prova vazia' % produto)
            print('   ATENCAO: nada comparado')
        wb.close()
        print()

    print('=' * 70)
    if falhas:
        print('FALHAS:')
        for f in falhas:
            print('   - %s' % f)
        sys.exit(1)
    print('EXCEL x CAMADA BI: SEM DIVERGENCIA NOS DOIS PRODUTOS')


if __name__ == '__main__':
    main()
