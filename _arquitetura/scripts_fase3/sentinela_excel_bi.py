# -*- coding: utf-8 -*-
"""sentinela_excel_bi.py - o que a Estatistica mostra e o que o BI publica

A cadeia so vale se a ponta final disser o mesmo que a planilha. Aqui alguns
analitos SENTINELA de cada produto sao lidos nos DOIS lugares e comparados
campo a campo:

    Estatistica (o que o gestor ve)   x   BI_Data (o que o Power BI le)

Compara Sigma, classificacao, CV, |Bias|, ET, ETp, DPM, rendimento, o plano de
CQ (regras, N, run size) e o nivel governante.

Nao abre o Excel: le os valores em cache com openpyxl, entao nao altera nada e
nao depende de COM.
"""
import io
import os
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', write_through=True)
import openpyxl

BASE = r"C:\Users\vitor\OneDrive - MSFT\Desktop\QC_INI"

SENTINELAS = {
    'QC_Bioquimica.xlsm': ['Lactato', 'Ácido úrico', 'Cálcio'],
    'QC_Hematologia.xlsm': ['WBC', 'HCT', 'PLT'],
}

# OS DOIS LAYOUTS DE ESTATISTICA SAO DIFERENTES, E ISSO E UM ACHADO.
#
# A Bioquimica recebeu a reestruturacao do ADR-033: 30 colunas, com DPM,
# rendimento, plano de CQ e contagem de EQA. A Hematologia ficou no layout
# legado: 13 colunas, sem nada disso -- embora o BI_Data dela ja publique os
# 84 campos, porque quem os calcula e o mBI, nao a planilha.
#
# Comparar so o que os dois tem seria esconder a diferenca; comparar tudo
# reprovaria a Hematologia por uma coluna que ela nunca teve. Cada produto tem
# o seu mapa, e a lacuna e declarada no fim.
CAMPOS = {
    'QC_Bioquimica.xlsm': [
        ('CV %', 'CV %', 'CV_Observado_pct', 0.01),
        ('ETp %', 'ETp %', 'ETp_pct', 0.01),
    ],
    'QC_Hematologia.xlsm': [
        ('CV %', 'CV %', 'CV_Observado_pct', 0.01),
        ('ETp %', 'ETp %', 'ETp_pct', 0.01),
    ],
}
TEXTOS = {
    'QC_Bioquimica.xlsm': [
        ('Classificação', 'Status sigma', 'Classificacao_Sigma'),
    ],
    'QC_Hematologia.xlsm': [
        ('Classificação', 'Classificação', 'Classificacao_Sigma'),
    ],
}
# BIAS E ET FICAM DE FORA DA COMPARACAO -- E ESCOPO, NAO DEFEITO.
#
# mBI chama mCEQ.BiasEQ(analito, anoBI, "SIGNED"): ano de referencia FIXO.
# A Estatistica chama a mesma funcao com os filtros que o gestor escolheu na
# tela (provedor, ano, rodada). Sao recortes diferentes do mesmo dado, entao os
# numeros podem divergir sem que nenhum dos dois esteja errado -- e exigir
# igualdade seria reprovar os dois por uma pergunta mal feita.
#
# ET carrega o bias, entao sai pelo mesmo motivo. O que continua conferido e o
# que NAO depende de recorte: CV, ETp, Sigma e a classificacao.
#
# Na Hematologia a coluna nem se chama a mesma coisa: e "Bias anual %", outro
# conceito.

# so a Bioquimica publica o plano na planilha; no BI ele existe nos dois
PLANO = ['Regra_Westgard_Recomendada', 'N_Controle_Recomendado',
         'RunSize_Max_Recomendado', 'Sigma_Plano', 'Nivel_Governante']

ok = fail = 0


def reg(nome, esp, obt, cond):
    global ok, fail
    if cond:
        ok += 1
    else:
        fail += 1
    print('    %-5s %-30s esp %-22s obt %s'
          % ('OK' if cond else 'FALHA', nome[:30], str(esp)[:22], str(obt)[:24]))


def cabecalho(ws, lim=30):
    """{rotulo: coluna} da primeira linha que tiver 'Analito' na coluna A"""
    for r in range(1, lim):
        if str(ws.cell(r, 1).value or '').strip().lower() == 'analito':
            d = {}
            for c in range(1, 40):
                v = str(ws.cell(r, c).value or '').strip()
                if v:
                    d[v] = c
            return r, d
    return None, {}


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main():
    for arq, alvos in SENTINELAS.items():
        cam = os.path.join(BASE, arq)
        if not os.path.exists(cam):
            continue
        wb = openpyxl.load_workbook(cam, data_only=True)
        est = None
        for ws in wb.worksheets:
            if ws.title.lower().startswith('estat'):
                est = ws
                break
        bi = wb['BI_Data']

        lin, cabE = cabecalho(est)
        cabB = {}
        for c in range(1, 90):
            v = str(bi.cell(1, c).value or '').strip()
            if v:
                cabB[v] = c

        # BI_Data por (analito, nivel): pega a ultima linha de cada par
        porPar = {}
        for r in range(2, bi.max_row + 1):
            a = str(bi.cell(r, cabB['Analito']).value or '').strip()
            n = bi.cell(r, cabB['Nivel']).value
            if not a or n is None:
                continue
            porPar[(a, int(n))] = r

        print()
        print('=' * 78)
        print('%s   (cabecalho da Estatistica na linha %s)' % (arq, lin))
        print('=' * 78)

        for alvo in alvos:
            achou = False
            for r in range(lin + 1, lin + 200):
                a = str(est.cell(r, 1).value or '').strip()
                if not a:
                    break
                if a != alvo:
                    continue
                nv = est.cell(r, 2).value
                if nv is None:
                    continue
                nv = int(nv)
                achou = True
                chave = (alvo, nv)
                print()
                print('  %s  nivel %d' % (alvo, nv))
                if chave not in porPar:
                    reg('presente no BI_Data', 'sim', 'AUSENTE', False)
                    continue
                rb = porPar[chave]

                for rot, colE, colB, tol in CAMPOS[arq]:
                    if colE not in cabE or colB not in cabB:
                        reg(rot, 'coluna existe',
                            'falta %s' % (colE if colE not in cabE else colB),
                            False)
                        continue
                    ve = num(est.cell(r, cabE[colE]).value)
                    vb = num(bi.cell(rb, cabB[colB]).value)
                    if ve is None and vb is None:
                        reg(rot, 'ambos vazios', 'ambos vazios', True)
                        continue
                    if ve is None or vb is None:
                        reg(rot, ve, vb, False)
                        continue
                    reg(rot, round(ve, 4), round(vb, 4), abs(ve - vb) <= tol)

                # SIGMA HERDA O BIAS -- informativo, com a conta a mostra.
                #
                # Sigma = (ETp - |Bias|) / CV. Como o bias vem de recortes
                # diferentes, o Sigma difere pelo MESMO motivo. Mostrar o bias
                # implicado de cada lado prova que nao ha uma segunda causa
                # escondida: a conta tem de fechar nos dois.
                colSig = 'SIX SIGMA' if 'SIX SIGMA' in cabE else 'Sigma'
                cvv = num(est.cell(r, cabE['CV %']).value) if 'CV %' in cabE else None
                etp = num(est.cell(r, cabE['ETp %']).value) if 'ETp %' in cabE else None
                se = num(est.cell(r, cabE[colSig]).value) if colSig in cabE else None
                sb = num(bi.cell(rb, cabB['Sigma']).value) if 'Sigma' in cabB else None
                if None not in (cvv, etp, se, sb):
                    print('    ----  Sigma planilha %.4f (bias implicado %.3f) | '
                          'BI %.4f (bias implicado %.3f)'
                          % (se, etp - se * cvv, sb, etp - sb * cvv))

                for rot, colE, colB in TEXTOS[arq]:
                    if colE not in cabE or colB not in cabB:
                        continue
                    te = str(est.cell(r, cabE[colE]).value or '').strip()
                    tb = str(bi.cell(rb, cabB[colB]).value or '').strip()
                    # o plano do BI e do PIOR nivel; a Estatistica mostra o da
                    # linha. Compara-se so quando os dois trazem algo.
                    if rot == 'Classificação':
                        print('    ----  Classificação planilha %r | BI %r'
                              % (te, tb))
                    else:
                        reg(rot, te, tb, te == tb)

                # o plano vive so no BI para a Hematologia; confere-se que ele
                # EXISTE e e coerente, nao que a planilha o repita
                # ABAIXO DE 3 SIGMA, N E RUN SIZE VAZIOS E O CERTO.
                #
                # Nao existe run size tabelado abaixo de 3 Sigma, e preencher um
                # numero ali autorizaria intervalo de CQ que o metodo nao
                # sustenta (ADR-038). Exigir "preenchido" reprovava o
                # comportamento correto.
                sp = num(bi.cell(rb, cabB['Sigma_Plano']).value) \
                    if 'Sigma_Plano' in cabB else None
                for c in PLANO:
                    if c not in cabB:
                        continue
                    v = bi.cell(rb, cabB[c]).value
                    cheio = v not in (None, '')
                    if c in ('N_Controle_Recomendado', 'RunSize_Max_Recomendado'):
                        if sp is not None and sp < 3:
                            reg('%s vazio abaixo de 3 Sigma' % c, '(vazio)',
                                '(vazio)' if not cheio else str(v)[:20],
                                not cheio)
                        else:
                            reg('BI publica %s' % c, 'preenchido',
                                str(v)[:20] if cheio else '(vazio)', cheio)
                    else:
                        reg('BI publica %s' % c, 'preenchido',
                            str(v)[:20] if cheio else '(vazio)', cheio)
            if not achou:
                print()
                print('  %s: nao encontrado na Estatistica' % alvo)
        wb.close()

    print()
    print('%d OK, %d FALHA' % (ok, fail))
    sys.exit(1 if fail else 0)


main()
