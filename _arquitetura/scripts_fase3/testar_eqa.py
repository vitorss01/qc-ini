# -*- coding: utf-8 -*-
"""testar_eqa.py - ADR-034: as provas do modulo de controle externo

Trabalha numa COPIA em TEMP. Nenhuma prova confia no que por acaso esta na
pasta: os numeros sao conferidos contra a planilha de referencia lida na hora,
ou contra conta feita a mao.

Uso: python testar_eqa.py <arquivo.xlsm> <referencia.xlsx>
"""
import io
import os
import re
import sys
import time
import shutil
import subprocess

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', write_through=True)
import win32com.client as w
import openpyxl

falhas = []


def ck(nome, cond, det=''):
    print(('  OK   ' if cond else '  FALHA') + '  ' + nome +
          (('  -> ' + det) if det else ''))
    if not cond:
        falhas.append(nome)


# Celula vazia volta do COM como None, e str(None) e 'None' -- string de quatro
# letras, que passa em qualquer teste de "nao vazio". Foi assim que a prova 7
# acusou "0 linhas sem analito canonico" enquanto o proprio motor contava 20.
# O erro era do instrumento, nao do produto.
def txt(v):
    return '' if v is None else str(v).strip()


def perto(a, b, tol=1e-6):
    try:
        return abs(float(a) - float(b)) <= tol
    except Exception:
        return False


def novo_excel():
    for t in range(1, 9):
        try:
            xl = w.DispatchEx('Excel.Application')
            xl.Visible = False
            xl.DisplayAlerts = False
            xl.EnableEvents = False
            xl.AutomationSecurity = 1
            return xl
        except Exception:
            if t in (1, 3, 5):
                subprocess.call(['powershell', '-NoProfile', '-Command',
                                 'Start-Process excel.exe -WindowStyle Minimized'],
                                stderr=subprocess.DEVNULL)
                time.sleep(10)
            time.sleep(2.5 * t)
    raise RuntimeError('Excel COM nao subiu')


def tenta(fn, vezes=10):
    ult = None
    for i in range(vezes):
        try:
            return fn()
        except Exception as e:
            ult = e
            s = str(e).lower()
            if 'rejeitada' not in s and 'rejected' not in s:
                raise
            time.sleep(1.0 + 0.8 * i)
    raise ult


def ler_ref(caminho):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb['CAP_Evaluation_Data']
    linhas = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 2).value:
            linhas.append([ws.cell(r, c).value for c in range(1, 17)])
    wb.close()
    return linhas


def main(caminho, referencia):
    ref = ler_ref(referencia)
    print('referencia: %d resultados' % len(ref))
    copia = os.path.join(os.environ.get('TEMP', '.'),
                         'teqa_' + os.path.basename(caminho))
    shutil.copy(caminho, copia)
    xl = novo_excel()
    wb = xl.Workbooks.Open(copia)
    try:
        base = wb.Worksheets('EQA_Base')
        base.Visible = -1
        cap = wb.Worksheets('EQA.CAP_Dados')
        ctl = wb.Worksheets('EQA.Controllab_Dados')
        res = wb.Worksheets('EQA.CAP_Resumo')
        nac = wb.Worksheets('EQA.CAP_Nao_Aceitaveis')
        es = wb.Worksheets('Estatística')
        xl.Calculation = -4105

        ultC = tenta(lambda: cap.Cells(cap.Rows.Count, 4).End(-4162).Row)
        d = tenta(lambda: cap.Range(cap.Cells(2, 1), cap.Cells(ultC, 18)).Value)
        print('EQA.CAP_Dados: %d linhas\n' % len(d))

        # ================================================================
        print('=== PROVA 1. as sete abas existem, na ordem, com a base oculta ===')
        ordem = [ws.Name for ws in wb.Worksheets]
        esperada = ['EQA.CAP_Dados', 'EQA.CAP_Resumo', 'EQA.CAP_Nao_Aceitaveis',
                    'EQA.Controllab_Dados', 'EQA.Controllab_Resumo',
                    'EQA.Controllab_Desvios', 'EQA_Base']
        i0 = ordem.index('EQA.CAP_Dados')
        ck('as 6 visiveis + a base estao agrupadas e na ordem',
           ordem[i0:i0 + 7] == esperada, ' > '.join(ordem[i0:i0 + 7]))
        wb2 = xl.Workbooks.Open(caminho) if False else None
        ck('tblEQA_CAP_Dados existe',
           any(lo.Name == 'tblEQA_CAP_Dados' for lo in cap.ListObjects))
        ck('tblEQA_Controllab_Dados existe',
           any(lo.Name == 'tblEQA_Controllab_Dados' for lo in ctl.ListObjects))
        ck('EQC_Dados preservada (nao apagada)',
           any(x == 'EQC_Dados' for x in ordem), 'presente entre as abas')

        print('\n=== PROVA 2. ordem dos dados: C-A inteiro, depois C-B, depois C-C ===')
        seq = []
        for row in d:
            s = str(row[1] or '')
            if s and (not seq or seq[-1] != s):
                seq.append(s)
        print('   sequencia de rodadas encontrada: %s' % seq)
        ck('cada rodada aparece num bloco unico (nada intercalado)',
           len(seq) == len(set(seq)), '%d blocos para %d rodadas distintas'
           % (len(seq), len(set(seq))))
        caps = [s for s in seq if s.startswith('C-')]
        ck('a ordem do CAP e C-A, C-B, C-C', caps == sorted(caps), str(caps))

        print('\n=== PROVA 3. Bias contra a conta a mao ===')
        # o exemplo da missao: resultado 93, media 90,4
        casos = 0
        for i, row in enumerate(d):
            r, m, bias = row[5], row[6], row[14]
            if isinstance(r, (int, float)) and isinstance(m, (int, float)) and m:
                esp = (r - m) / m * 100.0
                if not perto(bias, esp, 1e-9):
                    ck('linha %d: bias bate com ((X-M)/M)*100' % (i + 2), False,
                       '%s vs %s' % (bias, esp))
                    break
                casos += 1
        else:
            ck('as %d linhas com resultado e media tem bias exato' % casos, True)
        ck('exemplo da missao: 93 contra 90,4 = 2,876...',
           perto((93 - 90.4) / 90.4 * 100, 2.8761061946902655, 1e-9),
           '%.10f' % ((93 - 90.4) / 90.4 * 100))

        print('\n=== PROVA 4. |Bias| e magnitude, e o sinal fica preservado ===')
        neg = [row for row in d if isinstance(row[14], (int, float)) and row[14] < 0]
        ck('existem linhas com bias NEGATIVO (o sinal nao foi destruido)',
           len(neg) > 0, '%d linhas' % len(neg))
        ruim = [row for row in d
                if isinstance(row[14], (int, float))
                and not perto(row[15], abs(row[14]))]
        ck('|Bias| = ABS(Bias) em todas as linhas', not ruim,
           '%d divergencias' % len(ruim))
        if neg:
            ck('numa linha negativa, |Bias| ficou positivo',
               neg[0][15] > 0, 'bias=%.4f |bias|=%.4f' % (neg[0][14], neg[0][15]))

        print('\n=== PROVA 5. os numeros batem com a planilha de referencia ===')
        refUna = sum(1 for row in ref if str(row[9]).strip() == 'Unacceptable')
        refAcc = sum(1 for row in ref if str(row[9]).strip() == 'Acceptable')
        capUna = sum(1 for row in d if txt(row[11]) == 'Unacceptable')
        print('   referencia: %d Acceptable / %d Unacceptable' % (refAcc, refUna))
        ck('os %d Unacceptable da referencia chegaram inteiros' % refUna,
           capUna == refUna, '%d na pasta' % capUna)
        capRef = sum(1 for row in d
                     if str(row[17] or '').find('simula') < 0)
        ck('os %d resultados reais da referencia chegaram inteiros' % len(ref),
           capRef == len(ref), '%d linhas de origem real' % capRef)
        ck('a aba derivada mostra os mesmos %d Unacceptable' % refUna,
           tenta(lambda: nac.Cells(3, 1).Value).startswith(str(refUna)),
           repr(tenta(lambda: nac.Cells(3, 1).Value)))

        print('\n=== PROVA 6. o historico da EQC_Dados foi preservado, e isolado ===')
        sim = [row for row in d if str(row[17] or '').find('simula') >= 0]
        ck('os 90 registros da aba legada estao no modulo novo',
           len(sim) == 90, '%d linhas marcadas como simulacao' % len(sim))
        ultB = tenta(lambda: base.Cells(base.Rows.Count, 1).End(-4162).Row)
        b = tenta(lambda: base.Range(base.Cells(2, 1), base.Cells(ultB, 21)).Value)
        usoNao = sum(1 for row in b if txt(row[19]).upper() == 'NAO')
        usoSim = sum(1 for row in b if txt(row[19]).upper() == 'SIM')
        ck('na base, as 90 estao com Uso_Analitico = NAO', usoNao == 90, str(usoNao))
        ck('as %d reais estao com Uso_Analitico = SIM' % len(ref),
           usoSim == len(ref), str(usoSim))

        print('\n=== PROVA 7. EQA_Base consolidou os dois nomes de analito ===')
        ck('a base tem %d linhas' % len(d), (ultB - 1) == len(d),
           '%d' % (ultB - 1))
        ureia = [row for row in b if txt(row[3]) == 'Urea Nitrogen']
        ck('"Urea Nitrogen" preserva o nome do provedor na coluna D',
           len(ureia) > 0, '%d linhas' % len(ureia))
        ck('e recebe "Ureia" como analito canonico na coluna E',
           bool(ureia) and txt(ureia[0][4]) == 'Ureia',
           repr(ureia[0][4]) if ureia else '-')
        semCanon = [row for row in b if txt(row[3]) and not txt(row[4])]
        nomes = sorted(set(txt(row[3]) for row in semCanon))
        ck('so Ferritin e TSH ficaram sem correspondente',
           nomes == ['Ferritin', 'Thyroid Stim Hormone'], str(nomes))
        ck('e sao 20 linhas (15 de Ferritin + 5 de TSH)',
           len(semCanon) == 20, '%d' % len(semCanon))

        print('\n=== PROVA 8. chave logica: sem duplicidade silenciosa ===')
        chaves = [txt(row[20]) for row in b if txt(row[20])]
        ck('toda linha da base tem chave', len(chaves) == (ultB - 1),
           '%d chaves para %d linhas' % (len(chaves), ultB - 1))
        ck('nenhuma chave repetida', len(chaves) == len(set(chaves)),
           '%d repetidas' % (len(chaves) - len(set(chaves))))
        carimbo = str(tenta(lambda: base.Cells(1, 26).Value))
        ck('o carimbo relata as repetidas em vez de escondê-las',
           'chaves repetidas' in carimbo, carimbo[-30:])

        print('\n=== PROVA 9. status padronizado nao aprova por omissao ===')
        for termo, esperado in (('Acceptable', 'ACEITO'),
                                ('Unacceptable', 'NAO ACEITO'),
                                ('Satisfatorio', 'ACEITO'),
                                ('Insatisfatorio', 'NAO ACEITO'),
                                ('', 'NAO AVALIADO'),
                                ('bla bla', 'NAO AVALIADO')):
            got = str(tenta(lambda t=termo: xl.Run('PadronizarStatus', t)))
            ck('%r -> %s' % (termo, esperado), got == esperado, got)

        print('\n=== PROVA 10. o resumo e dinamico: nada fixado em 2025 ===')
        f = str(tenta(lambda: res.Cells(5, 1).Formula))
        ck('o rotulo da rodada vem de INDEX/MATCH, nao digitado',
           'INDEX' in f and 'MATCH' in f, f[:60])
        ck('nenhuma formula do resumo tem 2025 escrito',
           not any('2025' in str(tenta(lambda rr=r, cc=c: res.Cells(rr, cc).Formula))
                   for r in range(4, 30) for c in range(1, 11)))
        tot = tenta(lambda: res.Cells(13, 6).Value)
        ck('o TOTAL de Unacceptable do resumo = %d' % refUna,
           perto(tot, refUna), str(tot))

        print('\n=== PROVA 11. mCEQ passou a ler a EQA_Base ===')
        tenta(lambda: es.Range('L4').__setattr__('Value', 'CAP'))
        tenta(lambda: es.Range('N4').__setattr__('Value', 2025))
        tenta(lambda: es.Range('P4').__setattr__('Value', 'TODAS'))
        tenta(lambda: xl.CalculateFull())
        comBias = sum(1 for r in range(14, 94)
                      if isinstance(tenta(lambda rr=r: es.Cells(rr, 7).Value),
                                    (int, float)))
        ck('a Estatistica acendeu com o dado do CAP', comBias >= 50,
           '%d linhas com |Bias| numerico' % comBias)
        # 29 analitos mapeados x 2 niveis
        ck('sao 58 linhas (29 analitos com canonico x 2 niveis)',
           comBias == 58, str(comBias))

        print('\n=== PROVA 12. a simulacao NAO entra no bias ===')
        # Glicose tem linhas reais (C-A/B/C) e linhas de simulacao (A/B/C).
        reais = [row for row in b
                 if txt(row[4]) == 'Glicose'
                 and txt(row[19]).upper() == 'SIM'
                 and isinstance(row[16], (int, float))]
        simul = [row for row in b
                 if txt(row[4]) == 'Glicose'
                 and txt(row[19]).upper() == 'NAO'
                 and isinstance(row[16], (int, float))]
        mediaReais = sum(row[16] for row in reais) / len(reais) if reais else None
        mediaTudo = ((sum(row[16] for row in reais) + sum(row[16] for row in simul))
                     / (len(reais) + len(simul))) if (reais or simul) else None
        got = tenta(lambda: xl.Run('BiasEQ', 'Glicose', 2025, 'ABS', 'CAP', 'TODAS'))
        print('   |bias| medio so das reais = %.6f (%d linhas)'
              % (mediaReais, len(reais)))
        print('   |bias| medio se misturasse = %.6f (%d linhas)'
              % (mediaTudo, len(reais) + len(simul)))
        print('   BiasEQ devolveu             = %s' % got)
        ck('BiasEQ usa so as linhas reais', perto(got, mediaReais, 1e-9), str(got))
        ck('e NAO o valor contaminado pela simulacao',
           not perto(got, mediaTudo, 1e-9), '%.6f' % mediaTudo)

        print('\n=== PROVA 13. o filtro de rodada responde aos rotulos novos ===')
        vistos = {}
        for rod in ('C-A 2025', 'C-B 2025', 'C-C 2025'):
            v = tenta(lambda r=rod: xl.Run('BiasEQ', 'Glicose', 2025, 'ABS', 'CAP', r))
            vistos[rod] = v
            alvo = [row for row in b
                    if txt(row[4]) == 'Glicose'
                    and txt(row[2]) == rod
                    and isinstance(row[16], (int, float))]
            esp = sum(row[16] for row in alvo) / len(alvo) if alvo else None
            ck('%s -> media das %d amostras daquela rodada' % (rod, len(alvo)),
               perto(v, esp, 1e-9), '%s vs %s' % (v, esp))
        ck('as tres rodadas dao numeros diferentes',
           len(set(round(float(x), 8) for x in vistos.values())) == 3,
           str({k: round(float(v), 4) for k, v in vistos.items()}))

        print('\n=== PROVA 14. analito sem canonico nao alimenta a Estatistica ===')
        for nome in ('Ferritin', 'Thyroid Stim Hormone'):
            v = tenta(lambda n=nome: xl.Run('BiasEQ', n, 2025, 'ABS', 'CAP', 'TODAS'))
            ck('%s devolve SEM EP (e nao zero)' % nome, isinstance(v, str), repr(v))
        na = [str(es.Cells(r, 1).Value) for r in range(14, 94)]
        ck('nem aparece na lista de analitos da Estatistica',
           'Ferritin' not in na and 'Thyroid Stim Hormone' not in na)

        print('\n=== PROVA 15. linha nova calcula sozinha (UX da secao 31) ===')
        lo = cap.ListObjects('tblEQA_CAP_Dados')
        nAntes = lo.ListRows.Count
        nova = lo.ListRows.Add()
        r = nova.Range.Row
        for c, v in ((1, 'CAP'), (2, 'C-D 2026'), (3, 2026), (4, 'Glucose, serum'),
                     (5, 'CHM-99'), (6, 93.0), (7, 90.4), (12, 'Unacceptable')):
            tenta(lambda cc=c, vv=v: cap.Cells(r, cc).__setattr__('Value', vv))
        tenta(lambda: xl.CalculateFull())
        bias = tenta(lambda: cap.Cells(r, 15).Value)
        babs = tenta(lambda: cap.Cells(r, 16).Value)
        stat = tenta(lambda: cap.Cells(r, 23).Value)
        print('   linha nova %d: bias=%s |bias|=%s status=%r' % (r, bias, babs, stat))
        ck('a tabela cresceu sozinha', lo.ListRows.Count == nAntes + 1,
           '%d -> %d' % (nAntes, lo.ListRows.Count))
        ck('Bias calculou sem arrastar formula',
           perto(bias, (93 - 90.4) / 90.4 * 100, 1e-9), str(bias))
        ck('|Bias| calculou junto', perto(babs, abs(bias), 1e-9), str(babs))
        ck('status padronizado calculou junto', str(stat) == 'NAO ACEITO', repr(stat))

        print('\n=== PROVA 16. a derivada acompanha sem banco duplicado ===')
        antes = str(tenta(lambda: nac.Cells(3, 1).Value))
        tenta(lambda: xl.CalculateFull())
        depois = str(tenta(lambda: nac.Cells(3, 1).Value))
        ck('a lista de Unacceptable subiu de %s para %s' % (refUna, refUna + 1),
           depois.startswith(str(refUna + 1)), depois)
        achou = False
        for rr in range(5, 40):
            if str(tenta(lambda x=rr: nac.Cells(x, 5).Value)).strip() == 'CHM-99':
                achou = True
                break
        ck('a amostra nova apareceu na aba derivada', achou)

        print('\n=== PROVA 17. consolidar leva a linha nova ate a base ===')
        tenta(lambda: xl.Run('AtualizarEQABase'))
        ultB2 = tenta(lambda: base.Cells(base.Rows.Count, 1).End(-4162).Row)
        ck('a base cresceu junto', (ultB2 - 1) == (ultB - 1) + 1,
           '%d -> %d' % (ultB - 1, ultB2 - 1))
        carimbo2 = str(tenta(lambda: base.Cells(1, 26).Value))
        print('   %s' % carimbo2)
        ck('o carimbo registrou a consolidacao nova', 'total 546' in carimbo2,
           carimbo2[-60:])

        print('\n=== PROVA 18. QA de completude: nenhuma celula em erro ===')
        erros = []
        alvos = [('EQA.CAP_Dados', 60, 26), ('EQA.Controllab_Dados', 6, 26),
                 ('EQA.CAP_Resumo', 90, 12), ('EQA.Controllab_Resumo', 90, 12),
                 ('EQA.CAP_Nao_Aceitaveis', 40, 18),
                 ('EQA.Controllab_Desvios', 40, 18),
                 ('EQA_Base', 30, 26), ('Estatística', 40, 21)]
        for nome, r1, c1 in alvos:
            ws = wb.Worksheets(nome)
            for rr in range(1, r1 + 1):
                for cc in range(1, c1 + 1):
                    t = str(tenta(lambda a=ws, x=rr, y=cc: a.Cells(x, y).Text))
                    if t.startswith('#') and t.strip('#') != '':
                        erros.append('%s!%d,%d=%s' % (nome, rr, cc, t))
        ck('zero celulas em erro nas 8 areas varridas', not erros,
           '; '.join(erros[:5]))
    finally:
        try:
            wb.Close(False)
        except Exception:
            pass
        try:
            xl.Quit()
        except Exception:
            pass

    print('\n' + '=' * 68)
    if falhas:
        print('FALHAS (%d):' % len(falhas))
        for f in falhas:
            print('   - %s' % f)
        sys.exit(1)
    print('MODULO DE CONTROLE EXTERNO: TODAS AS PROVAS PASSARAM')


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
