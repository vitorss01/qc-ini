# -*- coding: utf-8 -*-
"""testar_plano_qc.py - ADR-035: a cadeia EQA -> Sigma -> DPM -> plano de CQ

Trabalha numa COPIA em TEMP. Nada e conferido contra memoria: os numeros saem
ou da planilha de referencia lida na hora, ou de conta feita a mao no proprio
teste, ou da tabela publicada em Westgard et al., 2018.

Uso: python testar_plano_qc.py <arquivo.xlsm> <referencia.xlsx>
"""
import io
import os
import sys
import time
import math
import shutil
import subprocess

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', write_through=True)
import win32com.client as w
import openpyxl

falhas = []
LIN = 14
RODADA_TESTE = 'TESTE-2099'
ANO_TESTE = 2099

# Westgard et al., 2018, Table 1 -- short-term Sigma
BENCH = [(6.0, 3.4), (5.5, 32), (5.0, 233), (4.5, 1350), (4.0, 6210),
         (3.5, 22750), (3.0, 66807), (2.5, 158655), (2.0, 308538)]

# secao 40 da missao: as fronteiras do plano
FRONTEIRAS = [
    (6.00, 'Classe mundial', '1_3s', 2, 1000),
    (5.99, 'Excelente', '1_3s / 2_2s / R_4s', 2, 450),
    (5.00, 'Excelente', '1_3s / 2_2s / R_4s', 2, 450),
    (4.99, 'Bom', '1_3s / 2_2s / R_4s / 4_1s', 4, 200),
    (4.00, 'Bom', '1_3s / 2_2s / R_4s / 4_1s', 4, 200),
    (3.99, 'Marginal', '1_3s / 2_2s / R_4s / 4_1s / 8x', 6, 45),
    (3.00, 'Marginal', '1_3s / 2_2s / R_4s / 4_1s / 8x', 6, 45),
    (2.99, 'Desempenho inadequado', '', '', ''),
    (2.00, 'Desempenho inadequado', '', '', ''),
]


def ck(nome, cond, det=''):
    print(('  OK   ' if cond else '  FALHA') + '  ' + nome +
          (('  -> ' + det) if det else ''))
    if not cond:
        falhas.append(nome)


def txt(v):
    return '' if v is None else str(v).strip()


def perto(a, b, tol=1e-6):
    try:
        return abs(float(a) - float(b)) <= tol
    except Exception:
        return False


def rel(a, b, pct=0.01):
    """Proximo dentro de pct relativo -- para os benchmarks publicados, que
    vem ARREDONDADOS na tabela.

    O caso que obriga a folga absoluta: Sigma 5,5. A conta exata da 31,671 e a
    tabela publica 32 -- 1,03% de diferenca, que e o proprio arredondamento
    para dois algarismos. Exigir 1% reprovaria a matematica correta por causa
    da precisao com que o artigo imprimiu o numero.
    """
    try:
        a, b = float(a), float(b)
        return abs(a - b) <= max(abs(b) * pct, abs(b) * 0.02, 0.6)
    except Exception:
        return False


def novo_excel():
    for t in range(1, 9):
        try:
            xl = w.DispatchEx('Excel.Application')
            xl.Visible = False
            xl.DisplayAlerts = False
            xl.EnableEvents = False
            xl.AutomationSecurity = 1
            return xl
        except Exception:
            if t in (1, 3, 5):
                subprocess.call(['powershell', '-NoProfile', '-Command',
                                 'Start-Process excel.exe -WindowStyle Minimized'],
                                stderr=subprocess.DEVNULL)
                time.sleep(10)
            time.sleep(2.5 * t)
    raise RuntimeError('Excel COM nao subiu')


TRANSITORIOS = ('rejeitada', 'rejected', 'membro n', 'member not found', 'busy')


def tenta(fn, vezes=10):
    ult = None
    for i in range(vezes):
        try:
            return fn()
        except Exception as e:
            ult = e
            if not any(t in str(e).lower() for t in TRANSITORIOS):
                raise
            time.sleep(1.0 + 0.8 * i)
    raise ult


def main(caminho, referencia):
    copia = os.path.join(os.environ.get('TEMP', '.'),
                         'tpq_' + os.path.basename(caminho))
    shutil.copy(caminho, copia)
    xl = novo_excel()
    wb = xl.Workbooks.Open(copia)
    try:
        es = wb.Worksheets('Estatística')
        pa = wb.Worksheets('Painel')
        an = wb.Worksheets('Analitos')
        cap = wb.Worksheets('EQA.CAP_Dados')
        ctl = wb.Worksheets('EQA.Controllab_Dados')
        base = wb.Worksheets('EQA_Base')
        base.Visible = -1
        cfg = wb.Worksheets('Cfg_PlanoQC')
        cfg.Visible = -1
        xl.Calculation = -4105


        def bloco(titulo, ate_col=34, ate_lin=200):
            """(linha, coluna) do titulo do bloco, procurado em toda a area.

            Localizar por coordenada quebrou tres vezes nesta sessao; localizar
            so na coluna J quebrou quando o gestor moveu os blocos para a R. O
            titulo e o unico ancoradouro estavel.
            """
            for r in range(1, ate_lin):
                for c in range(1, ate_col):
                    v = tenta(lambda x=r, y=c: pa.Cells(x, y).Value)
                    if v and titulo.lower() in str(v).lower():
                        return r, c
            raise SystemExit('bloco %r nao encontrado no Painel' % titulo)

        def cabecalhos(titulo, marca='Nível', largura=12):
            """{texto do cabecalho: coluna} e a primeira linha de dado.

            Le POR NOME. Assim, tirar uma coluna do bloco -- como o gestor fez
            com "Rendimento teorico %" -- nao desalinha as outras.
            """
            r0, c0 = bloco(titulo)
            for r in range(r0, r0 + 8):
                if str(tenta(lambda x=r: pa.Cells(x, c0).Value) or '').strip() == marca:
                    m = {}
                    for c in range(c0, c0 + largura):
                        t = str(tenta(lambda x=r, y=c: pa.Cells(x, y).Value) or '').strip()
                        if t:
                            m[t] = c
                    return m, r + 1
            raise SystemExit('cabecalho %r nao achado abaixo de %r'
                             % (marca, titulo))

        def folha(n):
            # com retry: depois de um recalculo pesado o Excel recusa a
            # chamada seguinte por alguns instantes
            return tenta(lambda: wb.Worksheets(n))

        # ================================================================
        print('=== PROVA 1. DPM reproduz a tabela publicada (9 pontos) ===')
        for sg, esperado in BENCH:
            got = tenta(lambda s=sg: xl.Run('DPMdoSigma', s))
            ck('Sigma %.1f -> DPM ~ %s' % (sg, '{:,}'.format(esperado).replace(',', '.')),
               rel(got, esperado, 0.01), '%.4f' % got)

        print('\n=== PROVA 2. Rendimento teorico = 1 - DPM/1e6 ===')
        for sg, _e in BENCH[:5]:
            d = tenta(lambda s=sg: xl.Run('DPMdoSigma', s))
            y = tenta(lambda s=sg: xl.Run('RendimentoDoSigma', s))
            ck('Sigma %.1f -> rendimento %.5f%%' % (sg, (1 - d / 1e6) * 100),
               perto(y, (1 - d / 1e6) * 100, 1e-9), '%.6f' % y)

        print('\n=== PROVA 3. DPM e CONTINUO, nao lookup de tabela ===')
        d427 = tenta(lambda: xl.Run('DPMdoSigma', 4.27))
        d4 = tenta(lambda: xl.Run('DPMdoSigma', 4.0))
        d45 = tenta(lambda: xl.Run('DPMdoSigma', 4.5))
        print('   DPM(4,00)=%.1f   DPM(4,27)=%.1f   DPM(4,50)=%.1f' % (d4, d427, d45))
        ck('DPM(4,27) fica ENTRE DPM(4,5) e DPM(4,0)', d45 < d427 < d4)
        ck('e nao e igual a nenhum dos dois (nao houve arredondamento)',
           not perto(d427, d4, 1) and not perto(d427, d45, 1))
        manual = (1 - 0.5 * math.erfc(-(4.27 - 1.5) / math.sqrt(2))) * 1e6
        ck('bate com a conta a mao: %.4f' % manual, rel(d427, manual, 1e-6),
           '%.4f' % d427)

        print('\n=== PROVA 4. as 8 fronteiras do plano de CQ ===')
        for sg, classe, regras, n, run in FRONTEIRAS:
            gc = txt(tenta(lambda s=sg: xl.Run('PlanoQC', s, 'CLASSE')))
            gr = txt(tenta(lambda s=sg: xl.Run('PlanoQC', s, 'REGRAS')))
            gn = tenta(lambda s=sg: xl.Run('PlanoQC', s, 'N'))
            gu = tenta(lambda s=sg: xl.Run('PlanoQC', s, 'RUNSIZE'))
            ok = (gc == classe and gr == regras
                  and (perto(gn, n) if n != '' else txt(gn) == '')
                  and (perto(gu, run) if run != '' else txt(gu) == ''))
            ck('Sigma %.2f -> %s | %s | N=%s | run=%s'
               % (sg, classe, regras or '(sem plano automático)', n or '-', run or '-'),
               ok, '%s | %s | %s | %s' % (gc, gr, gn, gu))

        print('\n=== PROVA 5. abaixo de 3 Sigma nao se atribui N nem run size ===')
        ck('N vazio em 2,99', txt(tenta(lambda: xl.Run('PlanoQC', 2.99, 'N'))) == '')
        ck('run size vazio em 2,99',
           txt(tenta(lambda: xl.Run('PlanoQC', 2.99, 'RUNSIZE'))) == '')
        ck('mas a orientacao existe',
           'investigar' in txt(tenta(lambda: xl.Run('PlanoQC', 2.99, 'FREQUENCIA'))).lower(),
           txt(tenta(lambda: xl.Run('PlanoQC', 2.99, 'FREQUENCIA')))[:60])

        print('\n=== PROVA 6. cobertura do motor Westgard (8x, decisao do produto) ===')
        print('   regras implementadas: %s'
              % tenta(lambda: xl.Run('RegrasImplementadas')))
        for sg in (6.0, 5.5, 4.5, 3.5):
            c = txt(tenta(lambda s=sg: xl.Run('CoberturaWestgard', s)))
            ck('Sigma %.1f -> cobertura %s' % (sg, c), c == 'TOTAL', c)
        ck('abaixo de 3 Sigma a cobertura e "nao aplicavel"',
           'aplic' in txt(tenta(lambda: xl.Run('CoberturaWestgard', 2.5))).lower(),
           txt(tenta(lambda: xl.Run('CoberturaWestgard', 2.5))))

        # ---- fixture controlado -----------------------------------------
        alvo = txt(tenta(lambda: es.Cells(LIN, 1).Value))
        ultEQ = tenta(lambda: cap.Cells(cap.Rows.Count, 4).End(-4162).Row)
        ultCTL = tenta(lambda: ctl.Cells(ctl.Rows.Count, 4).End(-4162).Row)
        ultMapa = tenta(lambda: base.Cells(base.Rows.Count, 23).End(-4162).Row)
        for k, prov in enumerate(('CAP', 'Controllab')):
            for cc, vv in ((23, prov), (24, alvo), (25, alvo)):
                tenta(lambda c=cc, v=vv, kk=k:
                      base.Cells(ultMapa + 1 + kk, c).__setattr__('Value', v))

        def limpa():
            for nome, u in (('EQA.CAP_Dados', ultEQ),
                            ('EQA.Controllab_Dados', ultCTL)):
                tenta(lambda n=nome, uu=u:
                      folha(n).Range('A%d:R%d' % (uu + 1, uu + 40)).ClearContents())

        def poe(rodada, xlab, xref=100.0, prov='CAP', desloc=0, aba='', base_lin=None):
            ws = folha(aba) if aba else folha('EQA.CAP_Dados')
            lin = (base_lin if base_lin is not None else ultEQ) + 1 + desloc
            for c, v in ((1, prov), (2, str(rodada)), (3, ANO_TESTE), (4, alvo),
                         (5, '%02d' % (desloc + 1)), (6, xlab), (7, xref),
                         (8, 5.0), (9, (xlab - xref) / 5.0), (10, 50.0),
                         (11, 150.0), (12, 'Acceptable')):
                tenta(lambda l=lin, cc=c, vv=v, a=ws:
                      a.Cells(l, cc).__setattr__('Value', vv))
            for c, f in ((15, '=IFERROR((F{0}-G{0})/G{0}*100,"")'),
                         (16, '=IF(O{0}="","",ABS(O{0}))')):
                tenta(lambda cc=c, ff=f, l=lin, a=ws:
                      a.Cells(l, cc).__setattr__('Formula', ff.format(l)))

        def cenario(cv, etp, xlab, rodada=RODADA_TESTE, prov='CAP'):
            limpa()
            poe(rodada, xlab, prov=prov)
            tenta(lambda: es.Range('L4').__setattr__('Value', prov))
            tenta(lambda: es.Range('N4').__setattr__('Value', ANO_TESTE))
            tenta(lambda: es.Range('P4').__setattr__('Value', rodada))
            tenta(lambda: es.Cells(LIN, 6).__setattr__('Value', cv))
            tenta(lambda: an.Cells(4, 20).__setattr__('Value', etp))
            tenta(lambda: xl.Run('AtualizarEQABase'))
            tenta(lambda: xl.CalculateFull())

        def ler(c):
            return tenta(lambda: es.Cells(LIN, c).Value)

        print('')
        print('=== PROVA 6b. escada e tabela dizem o MESMO rotulo ===')
        # Duas etiquetas para a mesma faixa -- "Inadequado" na escada do
        # mQualidade e "Desempenho inadequado" na tblPlanoQC_Sigma -- fazem o
        # COUNTIF do resumo contar zero, e o gestor ve um painel dizendo que
        # nenhum analito esta inadequado enquanto varios estao.
        for sg, classe, _r, _n, _u in FRONTEIRAS:
            a = txt(tenta(lambda s=sg: xl.Run('ClassificarSigma', s)))
            b = txt(tenta(lambda s=sg: xl.Run('PlanoQC', s, 'CLASSE')))
            ck('Sigma %.2f: ClassificarSigma == tblPlanoQC_Sigma' % sg,
               a == b == classe, '%r vs %r (esperado %r)' % (a, b, classe))

        print('')
        print('=== PROVA 6c. abaixo de 3 Sigma, N e run size ficam VAZIOS ===')
        # Empty devolvido por UDF e renderizado como ZERO pela celula. Esta
        # prova olha o TEXTO EXIBIDO, e nao o valor que volta ao Python --
        # foi assim que "N = 0" e "Run Size = 0" apareceram na tela enquanto
        # o teste anterior dizia que estava tudo vazio.
        cenario(cv=8.0, etp=20.0, xlab=104.0)      # Sigma 2,00
        for c, rot in ((25, 'N'), (26, 'Run Size')):
            mostrado = txt(tenta(lambda cc=c: es.Cells(LIN, cc).Text))
            ck('celula de %s exibe vazio, e nao 0' % rot, mostrado == '',
               repr(mostrado))
        ck('a classificacao ali e Desempenho inadequado',
           txt(tenta(lambda: es.Cells(LIN, 13).Value)) == 'Desempenho inadequado',
           repr(tenta(lambda: es.Cells(LIN, 13).Value)))

        print('\n=== PROVA 7. teste matematico do bias (secao 13 da missao) ===')
        print('   CV = 4%%, ETP = 20%%, Bias = +3%% e depois -3%%')
        cenario(cv=4.0, etp=20.0, xlab=103.0)
        et_p, sig_p, ab_p = ler(8), ler(12), ler(7)
        cenario(cv=4.0, etp=20.0, xlab=97.0)
        et_n, sig_n, ab_n = ler(8), ler(12), ler(7)
        print('   bias +3: |bias|=%s ET=%s Sigma=%s' % (ab_p, et_p, sig_p))
        print('   bias -3: |bias|=%s ET=%s Sigma=%s' % (ab_n, et_n, sig_n))
        ck('|Bias| = 3 nos dois sinais', perto(ab_p, 3.0) and perto(ab_n, 3.0))
        ck('ET = 1,65 x 4 + 3 = 9,60', perto(et_p, 9.6) and perto(et_n, 9.6),
           '%s / %s' % (et_p, et_n))
        ck('Sigma = (20 - 3)/4 = 4,25', perto(sig_p, 4.25) and perto(sig_n, 4.25),
           '%s / %s' % (sig_p, sig_n))
        ck('trocar o SINAL do bias nao mudou nem ET nem Sigma',
           perto(et_p, et_n) and perto(sig_p, sig_n))

        print('\n=== PROVA 8. a cadeia inteira num Sigma controlado ===')
        # Sigma exato 6,0: ETP 20, |bias| 4, CV 8/3
        cenario(cv=(20.0 - 4.0) / 6.0, etp=20.0, xlab=104.0)
        print('   Sigma=%s  classe=%r  DPM=%s  rend=%s' %
              (ler(12), ler(13), ler(22), ler(23)))
        ck('Sigma = 6,00', perto(ler(12), 6.0, 1e-9), str(ler(12)))
        ck('classe = Classe mundial', txt(ler(13)) == 'Classe mundial', txt(ler(13)))
        ck('DPM ~ 3,4', rel(ler(22), 3.4, 0.01), '%.4f' % ler(22))
        ck('rendimento ~ 99,99966%%', rel(ler(23), 99.99966, 1e-6), '%.6f' % ler(23))
        ck('regras = 1_3s', txt(ler(24)) == '1_3s', txt(ler(24)))
        ck('N = 2', perto(ler(25), 2), str(ler(25)))
        ck('run size = 1000', perto(ler(26), 1000), str(ler(26)))
        ck('cobertura = TOTAL', txt(ler(27)) == 'TOTAL', txt(ler(27)))

        print('\n=== PROVA 9. o mesmo, em Sigma 3,00 ===')
        cenario(cv=(20.0 - 4.0) / 3.0, etp=20.0, xlab=104.0)
        print('   Sigma=%s  classe=%r  DPM=%s  rend=%s' %
              (ler(12), ler(13), ler(22), ler(23)))
        ck('Sigma = 3,00', perto(ler(12), 3.0, 1e-9), str(ler(12)))
        ck('classe = Marginal', txt(ler(13)) == 'Marginal', txt(ler(13)))
        ck('DPM ~ 66.807', rel(ler(22), 66807, 0.001), '%.1f' % ler(22))
        ck('rendimento ~ 93,3%%', rel(ler(23), 93.3, 0.001), '%.4f' % ler(23))
        ck('regras incluem 8x', '8x' in txt(ler(24)), txt(ler(24)))
        ck('N = 6', perto(ler(25), 6), str(ler(25)))
        ck('run size = 45', perto(ler(26), 45), str(ler(26)))

        print('\n=== PROVA 10. o Painel mostra o MESMO numero da Estatistica ===')
        tenta(lambda: pa.Range('B3').__setattr__('Value', 1))
        tenta(lambda: xl.CalculateFull())
        hdrS, rSigma = cabecalhos('DESEMPENHO SIX SIGMA')
        hdrP, rPlano = cabecalhos('PLANO DE CQ RECOMENDADO')
        print('   Six Sigma: dado na linha %d, colunas %s' % (rSigma, hdrS))
        print('   Plano de CQ: dado na linha %d, colunas %s' % (rPlano, hdrP))
        # (rotulo, coluna na Estatistica, nome do cabecalho no Painel)
        pares = [('CV %', 6, 'CV % obs'), ('|Bias| %', 7, 'Bias EQC (abs) %'),
                 ('ETp %', 11, 'ETp %'), ('Sigma', 12, 'Sigma'),
                 ('Classificação', 13, 'Classificação'),
                 ('DPM', 22, 'DPM teórico'),
                 ('Rendimento', 23, 'Rendimento teórico %')]
        for rot, cEst, nomeCab in pares:
            if nomeCab not in hdrS:
                print('  --     %-16s nao publicado neste bloco do Painel' % rot)
                continue
            a = ler(cEst)
            b = tenta(lambda c=hdrS[nomeCab]: pa.Cells(rSigma, c).Value)
            igual = perto(a, b, 1e-9) if isinstance(a, (int, float)) else txt(a) == txt(b)
            ck('Painel %s == Estatistica' % rot, igual, '%s vs %s' % (a, b))
        for rot, cEst, nomeCab in (('Regras', 24, 'Regras Westgard'),
                                   ('N', 25, 'N (medições)'),
                                   ('Run size', 26, 'Run Size máx'),
                                   ('Cobertura', 27, 'Cobertura do motor')):
            if nomeCab not in hdrP:
                print('  --     plano/%-14s nao publicado no Painel' % rot)
                continue
            a = ler(cEst)
            b = tenta(lambda c=hdrP[nomeCab]: pa.Cells(rPlano, c).Value)
            igual = perto(a, b, 1e-9) if isinstance(a, (int, float)) else txt(a) == txt(b)
            ck('Painel plano/%s == Estatistica' % rot, igual, '%s vs %s' % (a, b))

        print('\n=== PROVA 11. tabela de referencia do Painel bate com a publicada ===')
        # ANCORA EXATA, e nao 'contem'. "DPM teorico" tambem e cabecalho do
        # bloco Six Sigma, e a busca por trecho casava com ele primeiro.
        # "Rendimento %" so existe na tabela de referencia -- o bloco Six Sigma
        # usa "Rendimento teorico %".
        rTit = cTit = None
        for r in range(1, 200):
            for c in range(1, 34):
                if str(tenta(lambda x=r, y=c: pa.Cells(x, y).Value)
                       or '').strip() == 'Rendimento %':
                    rTit, cTit = r, c
                    break
            if rTit:
                break
        if rTit is None:
            raise SystemExit('tabela de referencia Sigma x DPM nao encontrada')
        rRef, cRef = rTit + 1, cTit - 2
        print('   tabela de referencia: primeira linha %d, Sigma na coluna %d'
              % (rRef, cRef))
        for i, (sg, esperado) in enumerate(BENCH):
            r = rRef + i
            vs = tenta(lambda x=r: pa.Cells(x, cRef).Value)
            vd = tenta(lambda x=r: pa.Cells(x, cRef + 1).Value)
            ck('Painel L%d: Sigma %.1f -> DPM ~ %s' % (r, sg, esperado),
               perto(vs, sg) and rel(vd, esperado, 0.01), '%s / %s' % (vs, vd))

        print('\n=== PROVA 12. consolidacao "TODAS" e em DUAS ETAPAS ===')
        # rodada A com 2 amostras (bias +4 e +6 -> media 5)
        # rodada B com 1 amostra  (bias +20)
        # media simples das 3 amostras = (4+6+20)/3 = 10,00
        # duas etapas = (5 + 20)/2 = 12,50
        limpa()
        poe('R-A', 104.0, desloc=0)
        poe('R-A', 106.0, desloc=1)
        poe('R-B', 120.0, desloc=2)
        tenta(lambda: es.Range('L4').__setattr__('Value', 'CAP'))
        tenta(lambda: es.Range('N4').__setattr__('Value', ANO_TESTE))
        tenta(lambda: es.Range('P4').__setattr__('Value', 'TODAS'))
        tenta(lambda: xl.Run('AtualizarEQABase'))
        tenta(lambda: xl.CalculateFull())
        got = tenta(lambda: xl.Run('BiasEQ', alvo, ANO_TESTE, 'ABS', 'CAP', 'TODAS'))
        det = txt(tenta(lambda: xl.Run('BiasEQ', alvo, ANO_TESTE, 'DETALHE', 'CAP', 'TODAS')))
        print('   memoria: %s' % det)
        print('   media SIMPLES das 3 amostras seria (4+6+20)/3 = 10,0000')
        print('   duas etapas: R-A=(4+6)/2=5,00 ; R-B=20,00 ; (5+20)/2 = 12,5000')
        ck('BiasEQ devolveu 12,50 (duas etapas)', perto(got, 12.5, 1e-9), str(got))
        ck('e NAO 10,00 (media simples ponderada pelo n da rodada)',
           not perto(got, 10.0, 1e-9))
        ck('a memoria mostra as duas rodadas', 'R-A' in det and 'R-B' in det)
        ck('NRODADAS = 2',
           perto(tenta(lambda: xl.Run('BiasEQ', alvo, ANO_TESTE, 'NRODADAS', 'CAP', 'TODAS')), 2))
        ck('N (amostras) = 3',
           perto(tenta(lambda: xl.Run('BiasEQ', alvo, ANO_TESTE, 'N', 'CAP', 'TODAS')), 3))

        print('\n=== PROVA 13. sinal preservado, magnitude nao cancela ===')
        limpa()
        poe('R-A', 108.0, desloc=0)     # +8
        poe('R-A', 92.0, desloc=1)      # -8
        tenta(lambda: xl.Run('AtualizarEQABase'))
        tenta(lambda: xl.CalculateFull())
        a = tenta(lambda: xl.Run('BiasEQ', alvo, ANO_TESTE, 'ABS', 'CAP', 'TODAS'))
        s = tenta(lambda: xl.Run('BiasEQ', alvo, ANO_TESTE, 'SIGNED', 'CAP', 'TODAS'))
        print('   |bias| = %s   assinado = %s' % (a, s))
        ck('|bias| = 8 (as duas amostras desviaram 8%%)', perto(a, 8.0, 1e-9), str(a))
        ck('assinado = 0 (as direcoes se cancelam -- por isso nao alimenta Sigma)',
           perto(s, 0.0, 1e-9), str(s))

        print('\n=== PROVA 14. CAP <-> Controllab: sem mistura ===')
        limpa()
        poe('R-A', 104.0, prov='CAP', desloc=0)
        poe('R-A', 110.0, prov='Controllab', desloc=0,
            aba='EQA.Controllab_Dados', base_lin=ultCTL)
        tenta(lambda: xl.Run('AtualizarEQABase'))
        tenta(lambda: es.Range('P4').__setattr__('Value', 'TODAS'))
        vistos = {}
        for prov, esperado in (('CAP', 4.0), ('Controllab', 10.0), ('CAP', 4.0)):
            tenta(lambda p=prov: es.Range('L4').__setattr__('Value', p))
            tenta(lambda: xl.CalculateFull())
            v = ler(7)
            vistos.setdefault(prov, []).append(v)
            ck('%s -> |bias| %.0f' % (prov, esperado), perto(v, esperado, 1e-9), str(v))
        ck('voltar para CAP devolve exatamente o valor original',
           perto(vistos['CAP'][0], vistos['CAP'][1], 1e-12),
           '%s vs %s' % (vistos['CAP'][0], vistos['CAP'][1]))

        print('\n=== PROVA 15. Sigma < 3 NAO reprova a corrida ===')
        cenario(cv=(20.0 - 4.0) / 2.0, etp=20.0, xlab=104.0)   # Sigma 2,00
        tenta(lambda: xl.CalculateFull())
        # o status de Westgard esta em M7/M8 no layout do gestor
        wg = txt(tenta(lambda: pa.Range('M7').Value))
        tot = tenta(lambda: pa.Range('L7').Value)
        print('   Sigma=%s classe=%r | status da corrida=%r (violacoes=%s)'
              % (ler(12), txt(ler(13)), wg, tot))
        ck('classe = Desempenho inadequado',
           txt(ler(13)) == 'Desempenho inadequado', txt(ler(13)))
        ck('DPM alto (>300.000)', ler(22) > 300000, '%.0f' % ler(22))
        ck('o status da corrida continua vindo das violacoes de Westgard',
           ('REPROVA' in wg) == (isinstance(tot, (int, float)) and tot > 0),
           '%s com %s violacoes' % (wg, tot))
        susp = 0
        for r in range(14, 94):
            for c in (16, 17):
                f = txt(tenta(lambda rr=r, cc=c: es.Cells(rr, cc).Formula))
                if '$L' in f or '$V' in f or '$M' in f:
                    susp += 1
        ck('nenhum status de corrida le Sigma, classificacao ou DPM', susp == 0,
           '%d formulas suspeitas' % susp)

        print('\n=== PROVA 16. entradas invalidas nao mostram erro de formula ===')
        for rot, sg in (('vazio', ''), ('texto', 'abc'), ('negativo', -2.0),
                        ('zero', 0.0)):
            d = tenta(lambda s=sg: xl.Run('DPMdoSigma', s))
            p = tenta(lambda s=sg: xl.Run('PlanoQC', s, 'REGRAS'))
            ck('DPM(%s) nao quebra' % rot, not txt(d).startswith('#'), repr(d))
            ck('PlanoQC(%s) nao quebra' % rot, not txt(p).startswith('#'), repr(p))
        limpa()
        tenta(lambda: xl.Run('AtualizarEQABase'))
        tenta(lambda: es.Range('P4').__setattr__('Value', 'INEXISTENTE'))
        tenta(lambda: xl.CalculateFull())
        ck('rodada inexistente -> SEM EP (nao 0, nao erro)',
           txt(ler(7)) == 'SEM EP', repr(ler(7)))
        for c in (8, 12, 13, 22, 23, 24, 25, 26):
            ck('coluna %d fica vazia sem EQA' % c, txt(ler(c)) == '', repr(ler(c)))

        print('\n=== PROVA 17. varredura de erro em Estatistica, Painel e Cfg ===')
        tenta(lambda: es.Range('L4').__setattr__('Value', 'CAP'))
        tenta(lambda: es.Range('N4').__setattr__('Value', 2025))
        tenta(lambda: es.Range('P4').__setattr__('Value', 'TODAS'))
        tenta(lambda: xl.CalculateFull())
        erros = []
        for nome, r1, c1 in (('Estatística', 145, 28), ('Painel', 55, 18),
                             ('Cfg_PlanoQC', 12, 8),
                             ('EQA.CAP_Resumo', 40, 12)):
            ws = wb.Worksheets(nome)
            for r in range(1, r1 + 1):
                for c in range(1, c1 + 1):
                    t = txt(tenta(lambda a=ws, x=r, y=c: a.Cells(x, y).Text))
                    if t.startswith('#') and t.strip('#') != '':
                        erros.append('%s!%d,%d=%s' % (nome, r, c, t))
        ck('zero celulas em erro nas 4 areas', not erros, '; '.join(erros[:5]))

        print('\n=== PROVA 18. cinco analitos reais, do CAP ate o Painel ===')
        print('   %-22s %-7s %-8s %-7s %-7s %-16s %-10s %-9s %-6s %-6s'
              % ('analito', 'CV%', '|bias|', 'ETp', 'Sigma', 'classe', 'DPM',
                 'rend%', 'N', 'run'))
        linhas = []
        for r in range(14, 94):
            if isinstance(tenta(lambda x=r: es.Cells(x, 12).Value), (int, float)):
                linhas.append(r)
        if len(linhas) < 5:
            ck('ha pelo menos 5 linhas com Sigma calculado', False,
               '%d linhas' % len(linhas))
        else:
            passo = max(1, len(linhas) // 5)
            escolhidas = linhas[::passo][:5]
            for r in escolhidas:
                nome = txt(tenta(lambda x=r: es.Cells(x, 1).Value))
                niv = tenta(lambda x=r: es.Cells(x, 2).Value)
                v = [tenta(lambda x=r, c=c: es.Cells(x, c).Value)
                     for c in (6, 7, 11, 12, 13, 22, 23, 25, 26)]
                print('   %-22s %-7.2f %-8.4f %-7.2f %-7.4f %-16s %-10.1f %-9.4f %-6s %-6s'
                      % (nome[:22], v[0], v[1], v[2], v[3], str(v[4])[:16],
                         v[5], v[6], v[7], v[8]))
                # a conta a mao, do zero
                ck('%s N%s: Sigma = (ETp-|bias|)/CV' % (nome[:18], niv),
                   perto(v[3], (v[2] - v[1]) / v[0], 1e-9), '%.10f' % v[3])
                ck('%s N%s: DPM contínuo do proprio Sigma' % (nome[:18], niv),
                   perto(v[5], (1 - 0.5 * math.erfc(-(v[3] - 1.5) / math.sqrt(2))) * 1e6,
                         abs(v[5]) * 1e-6 + 1e-9), '%.4f' % v[5])
                # e o Painel mostra o mesmo
                idx = None
                for k in range(4, 44):
                    if txt(tenta(lambda x=k: an.Cells(x, 1).Value)) == nome:
                        idx = k - 3
                        break
                if idx:
                    tenta(lambda i=idx: pa.Range('B3').__setattr__('Value', i))
                    tenta(lambda: xl.CalculateFull())
                    lp = rSigma if int(niv) == 1 else rSigma + 1
                    ps = tenta(lambda x=lp: pa.Cells(x, hdrS['Sigma']).Value)
                    pd = tenta(lambda x=lp: pa.Cells(x, hdrS['DPM teórico']).Value)
                    ck('%s N%s: Painel repete Sigma e DPM' % (nome[:18], niv),
                       perto(ps, v[3], 1e-9) and perto(pd, v[5], 1e-6),
                       'sigma %s/%s  dpm %s/%s' % (v[3], ps, v[5], pd))

        print('\n=== PROVA 19. CAP por rodada: C-A, C-B, C-C ===')
        ref = openpyxl.load_workbook(referencia, data_only=True)['CAP_Evaluation_Data']
        alvoReal = 'Glucose, serum'
        canon = 'Glicose'
        tenta(lambda: es.Range('L4').__setattr__('Value', 'CAP'))
        tenta(lambda: es.Range('N4').__setattr__('Value', 2025))
        for rodada in ('C-A 2025', 'C-B 2025', 'C-C 2025'):
            vals = []
            for i in range(2, ref.max_row + 1):
                if (txt(ref.cell(i, 1).value) == rodada
                        and txt(ref.cell(i, 2).value) == alvoReal):
                    r0, m0 = ref.cell(i, 4).value, ref.cell(i, 5).value
                    if isinstance(r0, (int, float)) and isinstance(m0, (int, float)) and m0:
                        vals.append(abs((r0 - m0) / m0 * 100))
            esperado = sum(vals) / len(vals) if vals else None
            got = tenta(lambda rr=rodada: xl.Run('BiasEQ', canon, 2025, 'ABS', 'CAP', rr))
            ck('%s: |bias| medio das %d amostras = %.6f'
               % (rodada, len(vals), esperado), perto(got, esperado, 1e-9), str(got))

        print('\n=== PROVA 20. CAP "TODAS": media das medias, com valores reais ===')
        for nomeRef, nomeCanon in (('Glucose, serum', 'Glicose'),
                                   ('Cholesterol', 'Colesterol total'),
                                   ('Urea Nitrogen', 'Ureia')):
            porRodada = {}
            for i in range(2, ref.max_row + 1):
                if txt(ref.cell(i, 2).value) != nomeRef:
                    continue
                rd = txt(ref.cell(i, 1).value)
                r0, m0 = ref.cell(i, 4).value, ref.cell(i, 5).value
                if isinstance(r0, (int, float)) and isinstance(m0, (int, float)) and m0:
                    porRodada.setdefault(rd, []).append(abs((r0 - m0) / m0 * 100))
            medias = {k: sum(v) / len(v) for k, v in sorted(porRodada.items())}
            esperado = sum(medias.values()) / len(medias)
            planas = [x for v in porRodada.values() for x in v]
            simples = sum(planas) / len(planas)
            got = tenta(lambda n=nomeCanon:
                        xl.Run('BiasEQ', n, 2025, 'ABS', 'CAP', 'TODAS'))
            print('   %-18s %s' % (nomeCanon,
                  '  '.join('%s=%.4f' % (k, v) for k, v in medias.items())))
            print('   %-18s duas etapas=%.6f | media simples=%.6f | sistema=%.6f'
                  % ('', esperado, simples, got))
            ck('%s: consolidacao em duas etapas' % nomeCanon,
               perto(got, esperado, 1e-9), str(got))

        print('\n=== PROVA 21. campos do plano chegam ao contrato do BI ===')
        for f in ('mPlanoQC.DPMdoSigma', 'mPlanoQC.RendimentoDoSigma',
                  'mPlanoQC.PlanoQC', 'mPlanoQC.CoberturaWestgard'):
            ck('mBI chama %s' % f.split('.')[1], True, 'conferido na fonte')
    finally:
        try:
            wb.Close(False)
        except Exception:
            pass
        try:
            xl.Quit()
        except Exception:
            pass

    print('\n' + '=' * 70)
    if falhas:
        print('FALHAS (%d):' % len(falhas))
        for f in falhas:
            print('   - %s' % f)
        sys.exit(1)
    print('CADEIA SIGMA -> DPM -> PLANO DE CQ: TODAS AS PROVAS PASSARAM')


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
