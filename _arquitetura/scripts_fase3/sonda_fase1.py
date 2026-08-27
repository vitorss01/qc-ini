# -*- coding: utf-8 -*-
"""sonda_fase1.py - onde o banco da Fase 1 foi parar, e o que e $BT$1/$BU$1"""
import io
import os
import re
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', write_through=True)
import openpyxl

PRODUTOS = ['QC_Hematologia.xlsm', 'QC_Bioquimica.xlsm', 'QC_Imunologia.xlsm']
SCHEMA = ['RUN', 'Data', 'Nível', 'Lote', 'Analito', 'Resultado', 'Status', 'NC']

base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

for arq in PRODUTOS:
    cam = os.path.join(base, arq)
    if not os.path.exists(cam):
        continue
    wb = openpyxl.load_workbook(cam)
    print()
    print('=' * 78)
    print(arq)
    print('=' * 78)
    print('  abas com "result" no nome: %s'
          % [n for n in wb.sheetnames if 'result' in n.lower()])

    for nome in wb.sheetnames:
        if not nome.lower().startswith('db_result'):
            continue
        ws = wb[nome]
        cab = [str(ws.cell(1, c).value or '').strip() for c in range(1, 9)]
        print('  %s A1:H1 = %s' % (nome, cab))
        ok = all(a.lower() == b.lower() for a, b in zip(cab, SCHEMA))
        print('       schema da Fase 1: %s' % ('CONFERE' if ok else 'DIVERGE'))
        fmts = sorted({ws.cell(r, 1).number_format
                       for r in range(2, min(80, ws.max_row + 1))})
        print('       formato da coluna RUN: %s' % fmts)
        col_st = [str(ws.cell(r, 7).value or '') for r in range(2, ws.max_row + 1)]
        preench = [v for v in col_st if v]
        print('       Status: %s ativos, %s excluidos, %d linhas com valor'
              % (sum(1 for v in preench if v.lower().startswith('ativ')),
                 sum(1 for v in preench if v.lower().startswith('exclu')),
                 len(preench)))

    # o que sao BT1 e BU1 no Calc, e quem os usa
    if 'Calc' in wb.sheetnames:
        ws = wb['Calc']
        print('  Calc: BT1=%r  BU1=%r' % (ws['BT1'].value, ws['BU1'].value))
        print('        BS1=%r  BV1=%r' % (ws['BS1'].value, ws['BV1'].value))
        col = {}
        for linha in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200)):
            for c in linha:
                v = c.value
                if isinstance(v, str) and v.startswith('=') and re.search(r'\$B[TU]\$1\b', v):
                    col.setdefault(c.column_letter, 0)
                    col[c.column_letter] += 1
        if col:
            print('        colunas que citam $BT$1/$BU$1: %s' % col)
            for linha in ws.iter_rows(min_row=3, max_row=3):
                for c in linha:
                    v = c.value
                    if isinstance(v, str) and re.search(r'\$B[TU]\$1\b', str(v)):
                        print('        exemplo %s: %s' % (c.coordinate, str(v)[:150]))
                        break
                break
    # nomes definidos que poderiam substituir a referencia fixa
    nomes = [n for n in wb.defined_names if 'med' in n.lower() or 'dp' in n.lower()]
    print('  nomes definidos parecidos com media/DP: %s' % nomes[:8])
    wb.close()
