# -*- coding: utf-8 -*-
"""sonda_fase1c.py - #4 sem endereco fixo, coluna NC, e o flush do LiberStore

O achado #4 e sobre referencia absoluta fixa. Procurar por "$BT$1" seria
cometer o mesmo erro: na Bioquimica e na Imunologia o par mora em $AX$1/$AY$1.
Aqui a busca e pelo PADRAO -- qualquer $XX$1 usado como media/DP dentro das
formulas de Z-score do Calc -- e nao por um endereco.
"""
import io
import os
import re
import sys
from collections import Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', write_through=True)
import openpyxl
from oletools.olevba import VBA_Parser

PRODUTOS = ['QC_Hematologia.xlsm', 'QC_Bioquimica.xlsm', 'QC_Imunologia.xlsm']
base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
RX_ABS1 = re.compile(r'\$([A-Z]{1,3})\$1\b')

for arq in PRODUTOS:
    cam = os.path.join(base, arq)
    if not os.path.exists(cam):
        continue
    wb = openpyxl.load_workbook(cam)
    print()
    print('=' * 78)
    print(arq)
    print('=' * 78)

    # ---- #4: referencias absolutas a linha 1 dentro do Calc --------------
    ws = wb['Calc']
    cont = Counter()
    porcol = {}
    for linha in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200)):
        for c in linha:
            v = c.value
            if isinstance(v, str) and v.startswith('='):
                for m in RX_ABS1.finditer(v):
                    cont[m.group(0)] += 1
                    porcol.setdefault(m.group(0), set()).add(c.column_letter)
    print('  #4 referencias $XX$1 no Calc (top 8):')
    for ref, n in cont.most_common(8):
        alvo = ws[ref.replace('$', '')].value
        rot = str(alvo)[:52] if alvo is not None else '(vazia)'
        print('     %-8s %5d uso(s)  colunas %-28s conteudo: %s'
              % (ref, n, ','.join(sorted(porcol[ref])[:6]), rot))

    # ---- coluna NC (H) do banco -----------------------------------------
    db = wb['DB_Resultados']
    cabH = db.cell(3, 8).value
    amostraH = [db.cell(r, 8).value for r in range(4, 24)]
    naovazio = [v for v in amostraH if v not in (None, '')]
    print('  coluna H do DB_Resultados: cabecalho=%r, %d de 20 linhas com valor'
          % (cabH, len(naovazio)))
    larg = db.column_dimensions['H'].width if 'H' in db.column_dimensions else None
    print('     largura da coluna H: %s   (oculta: %s)'
          % (larg, db.column_dimensions['H'].hidden if 'H' in db.column_dimensions else '?'))

    wb.close()

    # ---- #5: quando o LiberStore e gravado -------------------------------
    p = VBA_Parser(cam)
    mods = {}
    for _, _, nome, cod in p.extract_macros():
        if nome:
            mods[nome.replace('.bas', '').replace('.cls', '')] = cod or ''
    p.close()

    print('  #5 gatilhos de gravacao das views por lote:')
    for nome in ('EstaPastaDeTrabalho', 'mLotes'):
        cod = mods.get(nome, '')
        for m in re.finditer(r'^\s*(Private |Public )?Sub\s+(\w+)', cod, re.M):
            corpo_ini = m.end()
            fim = cod.find('End Sub', corpo_ini)
            corpo = cod[corpo_ini:fim if fim > 0 else corpo_ini + 400]
            if re.search(r'FlushLoteAtual|SalvarViewNoBloco', corpo):
                print('     %s.%s -> chama o flush' % (nome, m.group(2)))
